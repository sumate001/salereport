import os
import re
import shutil
import tempfile
import zipfile
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─── Column indices (0-based) ─────────────────────────────────────────────────

class ItemCol:
    JOB              = 1
    DATE_PRINTED     = 2
    TITLE_TH         = 3
    PRICE            = 8
    ISBN             = 11
    TITLE_EN         = 14
    AGENCY           = 15
    ANNUAL_BI        = 17
    COPIES_PRINTED   = 18
    ROYALTY_RATE     = 32
    ADV              = 33
    ADV_CURRENCY     = 34
    PAYMENT_CURRENCY = 35
    BI_H1_AMARIN     = 39
    BI_H2_AMARIN     = 40
    BI_H1_ABOOK      = 41
    BI_H2_ABOOK      = 42
    ANNUAL_SOLD      = 43
    PREV_BALANCE     = 71
    STATUS_2024      = 72   # สถานะ 2024 (Annual) — ใช้ตรวจ balance_paid ปีก่อน
    STOCK_ACCOUNT    = 84
    EB_NET_BI1       = 74   # ยอดจ่าย 2025.1  (e-book H1 net receipt)
    EB_NET_BI2       = 77   # ยอดจ่าย 2025.2  (e-book H2 net receipt)
    EB_NET_ANNUAL    = 80   # ยอดจ่าย 2025    (e-book annual net receipt)
    STATUS_BI1       = 75   # สถานะ 2025.1
    STATUS_BI2       = 78   # สถานะ 2025.2
    STATUS_ANNUAL    = 81   # สถานะ 2025


# รหัสสินค้าบนปกเป็น EAN-13 เสมอ → ตรวจแค่ "13 หลัก" ห้ามผูกกับ prefix
#
# ของเดิมเขียนเป็น `978\d{10}` ซึ่งใช้ได้ตอนที่ input มีแต่หนังสือเล่มเดี่ยว พอชุดไฟล์ดิบ
# มี boxset / gift box ที่ใช้ EAN ไทย `885878…` (ในไฟล์ช่องนี้ยังชื่อ "ISBN" อยู่) ของ
# กลุ่มนั้นถูกทิ้งเงียบๆ ทั้งตอน map สัญญาและตอนวน item — prefix ใหม่ในอนาคตจะใช้ได้
# ทันทีโดยไม่ต้องแก้โค้ด ส่วนของที่หลุด pattern นี้ดูได้จาก scan_skipped_codes()
#
# ใช้ .match() เสมอ (ไม่ใช่ .search()) เพราะช่อง BookTH ของ intra เก็บเป็น
# "9789748491721 ชื่อหนังสือ" — ต้องตัดเฉพาะรหัสที่อยู่หน้าสุด
PRODUCT_CODE_RE = re.compile(r"'?(\d{13})")


def ean13_check_ok(code) -> bool:
    """ตรวจหลักตรวจสอบของ EAN-13

    ใช้เป็น **สัญญาณเตือน** เท่านั้น ห้ามเอามาคัดทิ้ง — พบรหัสที่ต้นทางพิมพ์ผิดค้างไว้
    เหมือนกันทั้งไฟล์ databook และ abook (เช่น 9786161843757 ที่มียอดขาย 357 เล่ม)
    ถ้าตัดทิ้งจะ join ไม่เจอทั้งที่ปัจจุบันทำงานได้ปกติเพราะสองฝั่งผิดตรงกัน
    """
    s = safe_str(code)
    if not re.fullmatch(r'\d{13}', s):
        return False
    digits = [int(c) for c in s]
    check = (10 - sum(d * (1 if i % 2 == 0 else 3) for i, d in enumerate(digits[:12])) % 10) % 10
    return check == digits[12]


# Canonical column order for the intra frame. Every intra file is re-indexed onto
# this layout by header name at load time (see ReportEngine.intra), so IntraCol
# indices stay valid no matter how the source file orders or renames its columns.
#
# ⚠ ไฟล์ RptRightAcc_* เปลี่ยน layout ทุกปี — ปี 2026 ย้าย Title(Eng Trans) จาก
# ตำแหน่ง 9 ไปท้ายไฟล์ ทำให้ทุกคอลัมน์ตั้งแต่ 9 เลื่อนซ้าย 1 ช่อง ถ้าอ่านด้วย index
# ตายตัวจะได้ royalty rate ผิดแบบไม่มี error
INTRA_CANONICAL = [
    'RightId', 'Paidtype', 'Agent(รับ report)', 'All Agents', 'All Proprietors',
    'Publisher', 'Country', 'ชื่อหนังสือตามใบขออนุมัติ', 'ชื่อในสัญญา',
    'Title(Eng Trans)', 'Title(Eng)', 'ชื่อไทย_Book', 'ชื่อไทย_EBook',
    'ชื่อไทย_EChapter', 'AdvPay', 'Royalty',
    'rt1_text', 'rt1_price', 'rt2_text', 'rt2_price', 'rt3_text', 'rt3_price',
    'rt4_text', 'rt4_price', 'rt5_text', 'rt5_price',
    'คำนวณ report โดยคิดจาก THB', 'วันเริ่มสัญญา', 'วันหมดอายุ', 'SellOffPeriod',
    'FinalSalesDate', 'สำนักพิมพ์', 'ประเภทลิขสิทธิ์', 'Editor Decision', 'RightCode',
    'Advanceใช้ร่วมกับBook', 'Advanceใช้ร่วมกับE-Book', 'Advanceใช้ร่วมกับE-Chapter',
    'Advanceใช้ร่วมกับAudioBook', 'Advanceใช้ร่วมกับE-Library', 'AdvanceNote',
    'BookTH01', 'BookTH02', 'BookTH03', 'BookTH04', 'BookTH05',
    'BookTH06', 'BookTH07', 'BookTH08', 'BookTH09', 'BookTH10',
    'On Copies Sold/Print', 'วิธีคำนวณRoyalty E-Book',
]


class IntraCol:
    """Positions within the canonical intra frame (INTRA_CANONICAL)."""
    PAIDTYPE   = INTRA_CANONICAL.index('Paidtype')
    AGENT      = INTRA_CANONICAL.index('Agent(รับ report)')
    PUBLISHER  = INTRA_CANONICAL.index('Publisher')
    COUNTRY    = INTRA_CANONICAL.index('Country')
    TITLE_EN   = INTRA_CANONICAL.index('Title(Eng)')
    ADV_PAY    = INTRA_CANONICAL.index('AdvPay')
    ROYALTY    = INTRA_CANONICAL.index('Royalty')
    START_DATE = INTRA_CANONICAL.index('วันเริ่มสัญญา')
    EXP_DATE   = INTRA_CANONICAL.index('วันหมดอายุ')     # E-Book
    SELL_OFF   = INTRA_CANONICAL.index('SellOffPeriod')  # Book
    # Tiered royalty rate columns
    RT1_TEXT = INTRA_CANONICAL.index('rt1_text');  RT1_PRICE = RT1_TEXT + 1
    RT2_TEXT = INTRA_CANONICAL.index('rt2_text');  RT2_PRICE = RT2_TEXT + 1
    RT3_TEXT = INTRA_CANONICAL.index('rt3_text');  RT3_PRICE = RT3_TEXT + 1
    RT4_TEXT = INTRA_CANONICAL.index('rt4_text');  RT4_PRICE = RT4_TEXT + 1
    RT5_TEXT = INTRA_CANONICAL.index('rt5_text');  RT5_PRICE = RT5_TEXT + 1
    # ISBNs: BookTH01–BookTH10 (auto-detected by content, see _init_intra_cols)


def _normalize_intra_header(name) -> str:
    """Collapse whitespace/newlines so header names match across file revisions."""
    if name is None or (isinstance(name, float) and name != name):
        return ''
    return re.sub(r'\s+', '', str(name)).lower()


_INTRA_CANONICAL_LOOKUP = {_normalize_intra_header(h): i for i, h in enumerate(INTRA_CANONICAL)}


def read_intra_file(path) -> pd.DataFrame:
    """Read one RptRightAcc_* file and re-index its columns onto INTRA_CANONICAL.

    The header sits on row index 2; data starts on row 3. Columns whose header is
    not in INTRA_CANONICAL are dropped; canonical columns missing from the file
    come back empty. Falls back to positional read if the header is unrecognisable.
    """
    raw = pd.read_excel(path, sheet_name=0, header=None, dtype=object)
    if raw.empty:
        return pd.DataFrame(columns=range(len(INTRA_CANONICAL)))

    header = raw.iloc[2] if len(raw) > 2 else pd.Series(dtype=object)
    src_for_canon = {}
    for src_idx, name in enumerate(header):
        key = _normalize_intra_header(name)
        canon_idx = _INTRA_CANONICAL_LOOKUP.get(key)
        if canon_idx is not None and canon_idx not in src_for_canon:
            src_for_canon[canon_idx] = src_idx

    data = raw.iloc[3:].reset_index(drop=True)

    # Too few headers recognised → header row is not where we expect it. Keep the
    # positional layout so at least the auto-detected job/ISBN matching still works.
    if len(src_for_canon) < 10:
        return data

    out = pd.DataFrame(index=data.index, columns=range(len(INTRA_CANONICAL)), dtype=object)
    for canon_idx, src_idx in src_for_canon.items():
        out.iloc[:, canon_idx] = data.iloc[:, src_idx].values

    # Carry over any extra source columns (e.g. unmapped ISBN-bearing columns) so
    # content-based auto-detection keeps working.
    mapped = set(src_for_canon.values())
    extra = [c for c in range(data.shape[1]) if c not in mapped]
    if extra:
        tail = data.iloc[:, extra]
        tail.columns = range(len(INTRA_CANONICAL), len(INTRA_CANONICAL) + len(extra))
        out = pd.concat([out, tail], axis=1)
    return out


class ExchangeCol:
    """ตำแหน่งคอลัมน์ในไฟล์อัตราแลกเปลี่ยน

    ไฟล์จัดเป็นบล็อกละปี ปีละ 4 คอลัมน์ (Q1–Q4) เรียงต่อกันจากปี BASE_YEAR
    ที่คอลัมน์ BASE_COL — ปี 2025 = 29–32, ปี 2026 = 33–36
    """
    BASE_YEAR = 2018
    BASE_COL  = 1

    @classmethod
    def quarter(cls, year: int, q: int) -> int:
        """คืน index คอลัมน์ของไตรมาส q (1–4) ในปีที่ระบุ"""
        return cls.BASE_COL + (year - cls.BASE_YEAR) * 4 + (q - 1)


# ─── Helpers ──────────────────────────────────────────────────────────────────

def safe_float(val, default=0.0):
    try:
        v = float(val)
        return default if (v != v) else v
    except Exception:
        return default


def safe_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val != val:
        return ''
    return str(val).strip()


def normalize_rate(rate):
    r = safe_float(rate)
    return r / 100 if r > 1 else r


def intra_flat_rate(intra_row):
    """อัตราค่าลิขสิทธิ์แบบ flat จากคอลัมน์ `Royalty` ของ intra (ไฟล์จริง = คอลัมน์ O)

    ใช้เมื่อไม่มีอัตราในไฟล์ item และสัญญาไม่ได้กรอก rt tier ไว้ — ซึ่งเป็นกรณีปกติ
    ของชุดข้อมูลไฟล์ดิบ (2026.1: 2,884 สัญญากรอก Royalty แบบ flat / กรอก rt tier
    แค่ 293 จากทั้งหมด 3,190) ถ้าไม่ fallback มาที่นี่ ช่อง ROYALTY RATE จะว่าง
    แล้ว AMOUNT (THB) กลายเป็น 0 ทั้งรายงาน
    """
    if intra_row is None or IntraCol.ROYALTY >= len(intra_row):
        return 0.0
    return normalize_rate(intra_row.iloc[IntraCol.ROYALTY])


def format_date_printed(val):
    """แปลง date เป็น 'Mar-14' format (Mon-YY) ตามตัวอย่าง"""
    if val is None:
        return ''
    if isinstance(val, datetime):
        return val.strftime('%b-%y')
    if hasattr(val, 'strftime'):
        return val.strftime('%b-%y')
    s = safe_str(val)
    if not s:
        return ''
    try:
        dt = pd.to_datetime(s)
        if pd.notna(dt):
            return dt.strftime('%b-%y')
    except Exception:
        pass
    return s


def extract_year(val):
    """ดึงปี ค.ศ. จาก cell ที่อาจเป็น date, int, float หรือ string"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.year
    if hasattr(val, 'year'):
        return val.year
    s = safe_str(val)
    if not s:
        return None
    m = re.search(r'(\d{4})', s)
    if m:
        y = int(m.group(1))
        return y - 543 if y > 2400 else y
    return None


DEFAULT_REPORT_YEAR = 2025


def report_year_from_period(period: str, year: int = DEFAULT_REPORT_YEAR) -> int:
    """ปีของ report — ทุก period ในชุดข้อมูลเดียวกันอยู่ปีเดียวกัน"""
    return year


def period_end_label(period: str, year: int = DEFAULT_REPORT_YEAR) -> str:
    if period == 'bi1':
        return f'For the Period Ended June 30, {year}'
    return f'For the Period Ended December 31, {year}'


def _date_sort_key(val):
    """Return a sortable datetime from a DATE_PRINTED cell value (oldest first)."""
    if isinstance(val, datetime):
        return val
    if hasattr(val, 'year'):
        return datetime(val.year, getattr(val, 'month', 1), getattr(val, 'day', 1))
    s = safe_str(val)
    if not s:
        return datetime(1900, 1, 1)
    try:
        dt = pd.to_datetime(s)
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass
    return datetime(1900, 1, 1)

def book_canon_key(title: str) -> str:
    """Aggressive normalisation used only for grouping books in search.
    Strips volume numbers and reprint markers so all print-runs of the same
    title collapse to one search result.
    """
    s = strip_reprint_suffix(title)
    s = re.sub(r'\s*[\(（][พP]\.\s*\d+(?:-\d+)?[\)）]', '', s)  # (พ.1) (พ.1-2) (P.3)
    s = re.sub(r'\s*\(\d+(?:-\d+)?\)', '', s)                    # (1) (1-2)
    s = re.sub(r'\s+\d+\.\d+', '', s)                            # " 5.1"  " 1.2"
    s = re.sub(r'\s+พ\.?\s*\d+', '', s)                          # " พ.5"
    s = re.sub(r'\s*:\s*[Vv]ol(?:ume)?\s*\d+', '', s)            # ": Volume 1"
    s = re.sub(r'\s+[Vv]ol(?:ume)?\s*\d+', '', s)                # " Volume 2"
    s = re.sub(r'\s+\d+$', '', s)                                 # trailing " 2"
    return s.strip().lower()


def strip_reprint_suffix(title: str) -> str:
    """Strip reprint/edition indicators and English-annotation suffixes from title."""
    s = re.sub(r'\s*[\(（][พP]\.\s*\d+[\)）]', '', title)
    s = re.sub(r'\s*พิมพ์เพิ่ม.*', '', s)
    s = re.sub(r'\s*พิมพ์ครั้งที่\s*\d+.*', '', s)
    # Strip parenthesized blocks that start with a Latin letter (English annotations)
    s = re.sub(r'\s*\([A-Za-z][^)]*\)', '', s)
    # Strip trailing ASCII-only production notes (e.g. "D-Print") only when the title
    # contains Thai characters (so we don't truncate English-only titles like "A Dance with Dragons")
    if re.search(r'[฀-๿]', s):
        s = re.sub(r'\s+[A-Za-z][A-Za-z0-9\-]*$', '', s)
    return s.strip()


def parse_rt_tiers(intra_row):
    """Parse rt1–rt5 tier columns from Intra row.
    Returns list of (min_copy, max_copy, rate) — rate normalized 0–1.
    'a' alone (no digits) or rate=0 → skipped.
    """
    if intra_row is None:
        return []

    RT_COLS = [
        (IntraCol.RT1_TEXT, IntraCol.RT1_PRICE),
        (IntraCol.RT2_TEXT, IntraCol.RT2_PRICE),
        (IntraCol.RT3_TEXT, IntraCol.RT3_PRICE),
        (IntraCol.RT4_TEXT, IntraCol.RT4_PRICE),
        (IntraCol.RT5_TEXT, IntraCol.RT5_PRICE),
    ]

    tiers = []
    for text_col, price_col in RT_COLS:
        try:
            rt_text  = safe_str(intra_row.iloc[text_col]).strip()
            rt_price = safe_float(intra_row.iloc[price_col])
        except (IndexError, KeyError):
            break
        if not rt_text or rt_price == 0:
            continue
        # strip commas from numbers (e.g. "a1-20,000" → "a1-20000"), allow trailing text
        rt_clean = re.sub(r'(?<=\d),(?=\d)', '', rt_text)
        m_range = re.match(r'a(\d+)-(\d+)', rt_clean, re.IGNORECASE)
        m_from  = re.match(r'a(\d+)',        rt_clean, re.IGNORECASE)
        if m_range:
            tiers.append((int(m_range.group(1)), int(m_range.group(2)), normalize_rate(rt_price)))
        elif m_from:
            tiers.append((int(m_from.group(1)), float('inf'), normalize_rate(rt_price)))
    return tiers


def apply_rt_tiers(copies_sold, tiers):
    """Split copies_sold across tiers. Returns [(copies_in_tier, rate), ...] for tiers with copies > 0."""
    result = []
    for min_c, max_c, rate in tiers:
        c = max(0.0, min(copies_sold, max_c) - (min_c - 1))
        if c > 0:
            result.append((c, rate))
    return result


# ─── ReportEngine ─────────────────────────────────────────────────────────────

class ReportEngine:
    def __init__(self, item_path: str, intra_paths, exchange_path: str,
                 year: int = DEFAULT_REPORT_YEAR, item_frame=None):
        self.year          = int(year)
        self.item_path     = item_path
        self.intra_paths   = intra_paths if isinstance(intra_paths, list) else [intra_paths]
        self.exchange_path = exchange_path
        # item_frame: ใช้ item master ที่ประกอบมาแล้ว (ชุดข้อมูลตั้งแต่ 2026.1 ที่ไม่มี
        # ไฟล์ item ส่งมา — ดู item_builder.build_item_frame) แทนการอ่านจาก Excel
        self._item_df      = item_frame
        self._intra_df     = None
        self._rates        = {}
        self._intra_job_col   = -1
        self._intra_isbn_col  = -1
        self._intra_isbn_cols = []
        self._intra_cols_init = False
        self._flat_rates      = None   # ISBN → อัตรา flat จาก intra (ดู flat_rate_for_isbn)

    # ── Lazy loaders ──────────────────────────────────────────────────────────

    @property
    def item(self):
        if self._item_df is None:
            self._item_df = pd.read_excel(
                self.item_path, sheet_name='Item', skiprows=2, header=None, dtype=object
            )
        return self._item_df

    @property
    def intra(self):
        if self._intra_df is None:
            dfs = [read_intra_file(p) for p in self.intra_paths]
            self._intra_df = pd.concat(dfs, ignore_index=True)
        return self._intra_df

    def _ensure_rates(self):
        if self._rates:
            return
        df = pd.read_excel(
            self.exchange_path, sheet_name='อัตราแลกเปลี่ยน', header=None, dtype=object
        )
        for _, row in df.iloc[4:].iterrows():
            currency = safe_str(row.iloc[0])
            if not currency:
                continue
            quarters = {}
            for q in range(1, 5):
                col = ExchangeCol.quarter(self.year, q)
                quarters[f'Q{q}'] = (
                    safe_float(row.iloc[col], 1.0) if col < len(row) else 1.0
                )
            self._rates[currency] = quarters

    def get_rate(self, currency: str, period: str) -> float:
        self._ensure_rates()
        lookup  = 'JYP' if currency == 'JPY' else currency
        rates   = self._rates.get(lookup, {'Q1': 1.0, 'Q2': 1.0, 'Q3': 1.0, 'Q4': 1.0})
        quarter = {'bi1': 'Q2', 'bi2': 'Q4', 'annual': 'Q4'}.get(period, 'Q4')
        return rates.get(quarter, 1.0) or 1.0

    # ── Cascading filter options ───────────────────────────────────────────────

    def _clean(self, series):
        return series.fillna('').astype(str).str.strip()

    def _build_isbn_to_contract(self):
        """Build isbn→contract_idx map from intra file. Returns (contracts, isbn_to_contract)."""
        self._init_intra_cols()
        contracts = []
        isbn_to_contract = {}
        for _, intra_row in self.intra.iterrows():
            isbns = []
            for isbn_col in self._intra_isbn_cols:
                val = safe_str(intra_row.iloc[isbn_col]).strip()
                m = PRODUCT_CODE_RE.match(val)
                if m:
                    isbn_val = m.group(1)
                    if isbn_val not in isbns:
                        isbns.append(isbn_val)
            if not isbns:
                continue
            idx = len(contracts)
            contracts.append({'isbns': isbns})
            for isbn in isbns:
                if isbn not in isbn_to_contract:
                    isbn_to_contract[isbn] = idx
        return contracts, isbn_to_contract

    def get_books(self, q: str = '') -> list:
        """Return title-level [{isbns, title_th, title_en}] — one entry per canonical title.

        All print runs / ISBNs sharing the same stripped title are merged into one result,
        so the user never sees duplicate entries for the same book.
        """
        q_lower = q.lower().strip()
        groups       = {}   # canon_key -> {isbns, isbn_set, title_th, title_en}
        order        = []   # first-appearance order
        isbn_to_key  = {}   # isbn -> canon_key already assigned

        for _, row in self.item.iterrows():
            isbn = safe_str(row.iloc[ItemCol.ISBN])
            if not isbn or not PRODUCT_CODE_RE.match(isbn):
                continue
            title_th = strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_TH]))
            title_en = strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_EN]))
            key = book_canon_key(title_th) or book_canon_key(title_en)
            if not key:
                continue
            # If this ISBN was already put in another group, follow that group
            key = isbn_to_key.get(isbn, key)
            if key not in groups:
                groups[key] = {
                    'isbns':    [],
                    'isbn_set': set(),
                    'title_th': title_th,
                    'title_en': title_en,
                }
                order.append(key)
            g = groups[key]
            if isbn not in g['isbn_set']:
                g['isbns'].append(isbn)
                g['isbn_set'].add(isbn)
                isbn_to_key[isbn] = key

        results = []
        for canonical in order:
            g = groups[canonical]
            if q_lower:
                if (q_lower not in g['title_th'].lower()
                        and q_lower not in g['title_en'].lower()
                        and not any(q_lower in i for i in g['isbns'])):
                    continue
            results.append({
                'isbns':    g['isbns'],
                'title_th': g['title_th'],
                'title_en': g['title_en'],
            })
        return results

    def get_countries(self):
        vals = self._clean(self.intra.iloc[:, IntraCol.COUNTRY])
        return sorted([v for v in vals.unique() if v and v.lower() != 'nan'])

    def get_agencies(self, country=''):
        df = self.intra
        if country:
            df = df[self._clean(df.iloc[:, IntraCol.COUNTRY]) == country]
        vals = self._clean(df.iloc[:, IntraCol.AGENT])
        return sorted([v for v in vals.unique() if v and v.lower() != 'nan' and len(v) > 3])

    def get_agencies_from_item(self, q: str = '') -> list:
        """Return sorted list of agency names from the item file, optionally filtered by query."""
        q_lower = q.lower().strip()
        agency_series = self._clean(self.item.iloc[:, ItemCol.AGENCY])
        agencies = sorted([
            a for a in agency_series.unique()
            if a and len(a) > 2 and a.lower() != 'nan'
        ])
        if q_lower:
            agencies = [a for a in agencies if q_lower in a.lower()]
        return agencies

    def get_publishers(self, agency, country=''):
        df = self.intra
        df = df[self._clean(df.iloc[:, IntraCol.AGENT]) == agency]
        if country:
            df = df[self._clean(df.iloc[:, IntraCol.COUNTRY]) == country]
        vals = self._clean(df.iloc[:, IntraCol.PUBLISHER])
        return sorted([v for v in vals.unique() if v and v.lower() != 'nan'])

    # ── Sell-off filter ───────────────────────────────────────────────────────

    def _init_intra_cols(self):
        if self._intra_cols_init:
            return
        self._intra_cols_init = True
        # ต้องกวาดทุกแถว — ของเดิมดูแค่ 200 แถวแรก ทำให้ BookTH05–10 ไม่ถูกนับเป็น
        # คอลัมน์ ISBN (คอลัมน์ท้ายๆ มี ISBN แค่หลักสิบแถวและอยู่ลึกในไฟล์) เล่มที่ ISBN
        # ไปตกอยู่คอลัมน์เหล่านั้นจึงหาสัญญาไม่เจอ → ออกรายงานแบบ orphan ไม่มีอัตรา
        # ค่าลิขสิทธิ์/advance และ agent ผิด (ชุด 2026.1: 185 ISBN)
        for c in range(self.intra.shape[1]):
            col = self._clean(self.intra.iloc[:, c])
            if col.str.match(r'\d{2}/[A-Z]').any() or col.str.startswith('EB/').any():
                self._intra_job_col = c
            if col.str.match(r"'?\d{13}").any():
                self._intra_isbn_cols.append(c)
        self._intra_isbn_col = self._intra_isbn_cols[0] if self._intra_isbn_cols else -1

    def flat_rate_for_isbn(self, isbn: str) -> float:
        """อัตรา flat (`Royalty`) ต่อ ISBN — สำหรับที่ที่เรียก per-row จนใช้ _find_intra_row ไม่ไหว

        dashboard วนทุกแถวของ item (หลักหมื่น) การหา intra row ทีละแถวช้าเกินไป
        จึงกวาด intra รอบเดียวแล้ว cache ไว้ ใช้ rt tier ก่อนถ้าสัญญามีกรอกไว้
        """
        if self._flat_rates is None:
            self._init_intra_cols()
            rates = {}
            for _, row in self.intra.iterrows():
                tiers = parse_rt_tiers(row)
                rate  = tiers[0][2] if tiers else intra_flat_rate(row)
                if not rate:
                    continue
                for c in self._intra_isbn_cols:
                    m = PRODUCT_CODE_RE.match(safe_str(row.iloc[c]))
                    if m and m.group(1) not in rates:
                        rates[m.group(1)] = rate
            self._flat_rates = rates
        return self._flat_rates.get(safe_str(isbn), 0.0)

    def intra_codes(self) -> set:
        """รหัสสินค้าทุกตัวที่ลงทะเบียนไว้ในไฟล์ intra (คอลัมน์ BookTH01–10)"""
        self._init_intra_cols()
        codes = set()
        for c in self._intra_isbn_cols:
            for v in self.intra.iloc[:, c]:
                m = PRODUCT_CODE_RE.match(safe_str(v).strip())
                if m:
                    codes.add(m.group(1))
        return codes

    def scan_skipped_codes(self, period: str = 'bi1', sample: int = 15) -> dict:
        """รายงาน "รหัสที่ระบบข้าม" — ของที่หลุดจากรายงาน พร้อมเหตุผลและจำนวนเล่มที่ขายได้

        มีไว้ให้ข้อมูลแปลกๆ โผล่ขึ้นหน้าจอเองตั้งแต่ตอนอัปโหลด แทนที่จะรอให้เจ้าของ
        ลิขสิทธิ์ทักมาว่ายอดไม่ตรง (เคสจริง: boxset ที่ใช้ EAN `885878…` หายทั้งกลุ่ม
        โดยไม่มีอะไรเตือน) 4 กอง เรียงตามความรุนแรง

          bad_format   รหัสไม่ใช่ 13 หลัก → `PRODUCT_CODE_RE` ไม่รับ ระบบข้ามแน่นอน
          no_agency    มีสัญญาใน intra + มียอดขาย แต่ databook ไม่ได้กรอก Agency →
                       เล่มไปกอง 'Direct Publisher' แทน agent จริง และถ้าไม่ได้กรอก
                       Annual/Bi-Annual ด้วยจะไม่ขึ้นรอบ bi1/bi2 เลย
          no_contract  รหัสใช้ได้ + มียอดขาย แต่ไม่มีในไฟล์ intra → ไม่มีสัญญาให้อ้างอิง
                       (ขึ้นรายงานเป็นไฟล์ orphan แต่ไม่มีอัตราค่าลิขสิทธิ์/advance)
          check_digit  13 หลักแต่หลักตรวจสอบ EAN-13 ไม่ผ่าน → ยังทำรายงานได้ตามปกติ
                       เป็นแค่สัญญาณว่าต้นทางน่าจะพิมพ์ผิด (ดู ean13_check_ok)

        เล่มที่ไม่มีทั้ง Agency และสัญญาถูกข้ามไป — ไฟล์ databook มีหนังสือของ Amarin เอง
        ปนมาสองในสามของแถว ถ้านับด้วยรายงานนี้จะกลายเป็นรายการหนังสือทั่วไปหลักพัน
        """
        sold_cols = {
            'bi1':    (ItemCol.BI_H1_AMARIN, ItemCol.BI_H1_ABOOK),
            'bi2':    (ItemCol.BI_H2_AMARIN, ItemCol.BI_H2_ABOOK),
            'annual': (ItemCol.ANNUAL_SOLD,),
        }.get(period, (ItemCol.BI_H1_AMARIN, ItemCol.BI_H1_ABOOK))

        known   = self.intra_codes()
        order   = ('bad_format', 'no_agency', 'no_contract', 'check_digit')
        buckets = {k: {'rows': [], 'codes': set()} for k in order}

        for _, row in self.item.iterrows():
            code   = safe_str(row.iloc[ItemCol.ISBN]).strip()
            agency = safe_str(row.iloc[ItemCol.AGENCY])
            if agency.lower() == 'nan':
                agency = ''
            sold   = sum(safe_float(row.iloc[c]) for c in sold_cols)
            m      = PRODUCT_CODE_RE.match(code)

            if not m:
                if not agency:
                    continue
                key = 'bad_format'
            else:
                code = m.group(1)
                has_contract = code in known
                if not agency:
                    if not (has_contract and sold > 0):
                        continue
                    key = 'no_agency'
                elif sold > 0 and not has_contract:
                    key = 'no_contract'
                elif not ean13_check_ok(code):
                    key = 'check_digit'
                else:
                    continue

            buckets[key]['codes'].add(code)
            buckets[key]['rows'].append({
                'job':         safe_str(row.iloc[ItemCol.JOB]),
                'code':        code,
                'title':       (strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_TH]))
                                or strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_EN]))),
                'agency':      agency,
                # รอบของสัญญา — ไฟล์ bi1/bi2 ออกเฉพาะ BI ส่วน annual ออกที่เหลือ
                # ตัวเลขนี้บอกว่ารายการนั้นจะไปโผล่ตอน generate รอบไหน
                'paidtype':    safe_str(row.iloc[ItemCol.ANNUAL_BI]),
                'copies_sold': int(sold),
            })

        labels = {
            'bad_format':  ('รหัสไม่ใช่ 13 หลัก', 'ระบบข้ามทิ้ง ไม่ขึ้นรายงาน', 'drop'),
            'no_agency':   ('มีสัญญาแต่ databook ไม่ได้กรอก Agency',
                            'ไปกอง Direct Publisher / ไม่ขึ้นรอบ BI', 'drop'),
            'no_contract': ('ไม่มีสัญญาในไฟล์ intra', 'ขึ้นรายงานแต่ไม่มีอัตราค่าลิขสิทธิ์', 'warn'),
            'check_digit': ('หลักตรวจสอบ EAN-13 ไม่ผ่าน', 'ยังใช้งานได้ — น่าจะพิมพ์ผิด', 'info'),
        }
        return {
            'period':     period,
            'item_rows':  int(len(self.item)),
            'intra_codes': len(known),
            'buckets': [
                {
                    'id':          key,
                    'label':       labels[key][0],
                    'effect':      labels[key][1],
                    'severity':    labels[key][2],
                    'codes':       len(b['codes']),
                    'rows':        len(b['rows']),
                    'copies_sold': int(sum(r['copies_sold'] for r in b['rows'])),
                    'samples':     sorted(b['rows'], key=lambda r: -r['copies_sold'])[:sample],
                }
                for key, b in ((k, buckets[k]) for k in order)
            ],
        }

    def _find_intra_row(self, item_row):
        self._init_intra_cols()
        isbn = safe_str(item_row.iloc[ItemCol.ISBN])
        job  = safe_str(item_row.iloc[ItemCol.JOB])
        df   = self.intra

        def _best(matched_df):
            """Return row with rt rate data if available, else first row."""
            for _, row in matched_df.iterrows():
                try:
                    if safe_float(row.iloc[IntraCol.RT1_PRICE]) > 0:
                        return row
                except Exception:
                    pass
            return matched_df.iloc[0]

        if self._intra_job_col >= 0 and job:
            mask = self._clean(df.iloc[:, self._intra_job_col]) == job
            if mask.any():
                return _best(df[mask])

        if isbn:
            for isbn_col in self._intra_isbn_cols:
                mask = self._clean(df.iloc[:, isbn_col]).str.startswith(isbn)
                if mask.any():
                    return _best(df[mask])

        return None

    def _passes_selloff(self, item_row, period: str) -> bool:
        report_yr  = report_year_from_period(period, self.year)
        job        = safe_str(item_row.iloc[ItemCol.JOB])
        is_ebook   = job.upper().startswith('EB/')
        self._init_intra_cols()
        isbn = safe_str(item_row.iloc[ItemCol.ISBN])
        df   = self.intra
        col  = IntraCol.EXP_DATE if is_ebook else IntraCol.SELL_OFF

        # Collect all matching intra rows (by job or ISBN) to get the most permissive date.
        # A single contract can have multiple sub-rows; if ANY row allows the book (sell_off
        # is None or >= report_year), include it.
        def _any_pass(matched_df):
            for _, row in matched_df.iterrows():
                sell_yr = extract_year(row.iloc[col])
                if sell_yr is None or sell_yr >= report_yr:
                    return True
            return False

        if self._intra_job_col >= 0 and job:
            mask = self._clean(df.iloc[:, self._intra_job_col]) == job
            if mask.any():
                return _any_pass(df[mask])

        if isbn:
            for isbn_col in self._intra_isbn_cols:
                mask = self._clean(df.iloc[:, isbn_col]).str.startswith(isbn)
                if mask.any():
                    return _any_pass(df[mask])

        return True

    # ── Row builder ───────────────────────────────────────────────────────────

    def _build_rows(self, items, period: str):
        rows = []
        if items.empty:
            return rows

        # Group by ISBN (maintain first-appearance order)
        isbn_to_group = {}
        isbn_order    = []
        for _, r in items.iterrows():
            isbn = safe_str(r.iloc[ItemCol.ISBN])
            if isbn not in isbn_to_group:
                isbn_to_group[isbn] = []
                isbn_order.append(isbn)
            isbn_to_group[isbn].append(r)

        adv_placed = False  # advance/balance shown once per contract, not per ISBN

        for isbn in isbn_order:
            group = isbn_to_group[isbn]

            if not any(self._passes_selloff(r, period) for r in group):
                continue

            first_r       = group[0]
            is_ebook      = safe_str(first_r.iloc[ItemCol.JOB]).upper().startswith('EB/')
            retail        = safe_float(first_r.iloc[ItemCol.PRICE])
            fallback_rate = normalize_rate(first_r.iloc[ItemCol.ROYALTY_RATE])
            currency      = safe_str(first_r.iloc[ItemCol.PAYMENT_CURRENCY]) or 'USD'
            adv           = safe_float(first_r.iloc[ItemCol.ADV])
            adv_cur       = safe_str(first_r.iloc[ItemCol.ADV_CURRENCY])
            amt_cur       = adv_cur or currency
            status_2024   = safe_str(first_r.iloc[ItemCol.STATUS_2024])
            prev_balance  = safe_float(first_r.iloc[ItemCol.PREV_BALANCE])
            balance_paid  = abs(safe_float(first_r.iloc[ItemCol.PREV_BALANCE])) if status_2024 == 'จ่ายแล้ว' else 0.0
            _status_col   = {'bi1': ItemCol.STATUS_BI1, 'bi2': ItemCol.STATUS_BI2}.get(period, ItemCol.STATUS_ANNUAL)
            pay_status    = safe_str(first_r.iloc[_status_col])
            stock_account = safe_float(first_r.iloc[ItemCol.STOCK_ACCOUNT])
            ex_rate       = self.get_rate(currency, period)
            amt_ex_rate   = self.get_rate(amt_cur, period) if amt_cur != currency else ex_rate
            intra_row     = self._find_intra_row(first_r)
            tiers         = parse_rt_tiers(intra_row)
            # If item file has no rate but intra tiers do, use first tier rate as fallback
            if not fallback_rate and tiers:
                fallback_rate = tiers[0][2]
            # ยังไม่ได้อัตราอีก → ใช้ `Royalty` แบบ flat ของ intra เป็นด่านสุดท้าย
            if not fallback_rate:
                fallback_rate = intra_flat_rate(intra_row)

            # E-book net receipt (stored as negative; abs = royalty amount in payment currency)
            if is_ebook:
                if period == 'bi1':
                    eb_net = abs(safe_float(first_r.iloc[ItemCol.EB_NET_BI1]))
                elif period == 'bi2':
                    eb_net = abs(safe_float(first_r.iloc[ItemCol.EB_NET_BI2]))
                else:
                    eb_net = abs(safe_float(first_r.iloc[ItemCol.EB_NET_ANNUAL]))
            else:
                eb_net = 0.0

            # Clean titles (strip reprint suffix) — for display header
            title_th_raw  = safe_str(first_r.iloc[ItemCol.TITLE_TH])
            title_en_raw  = safe_str(first_r.iloc[ItemCol.TITLE_EN])
            clean_th      = strip_reprint_suffix(title_th_raw)
            clean_en      = strip_reprint_suffix(title_en_raw)
            title_display = clean_en or clean_th
            title_th_sub  = clean_th if clean_en else ''

            # Per-run copies_sold and totals
            run_data             = []
            total_copies_sold    = 0.0
            total_copies_printed = 0.0
            for r in group:
                if period == 'bi1':
                    run_sold = safe_float(r.iloc[ItemCol.BI_H1_AMARIN]) + safe_float(r.iloc[ItemCol.BI_H1_ABOOK])
                elif period == 'bi2':
                    run_sold = safe_float(r.iloc[ItemCol.BI_H2_AMARIN]) + safe_float(r.iloc[ItemCol.BI_H2_ABOOK])
                else:
                    run_sold = safe_float(r.iloc[ItemCol.ANNUAL_SOLD])
                run_data.append((r, run_sold))
                total_copies_sold    += run_sold
                total_copies_printed += safe_float(r.iloc[ItemCol.COPIES_PRINTED])

            # Sort print runs by date printed (oldest first)
            run_data.sort(key=lambda x: _date_sort_key(x[0].iloc[ItemCol.DATE_PRINTED]))

            # Tier split: determined by each run's position in the cumulative
            # print sequence.  A run occupying [cumul+1 … cumul+printed] may
            # straddle multiple tier brackets; its sold copies are split
            # proportionally.  run_portions[i] = [(p_printed, tier_idx, rate), …]
            if tiers:
                cumul_printed = 0.0
                tier_buckets  = [0.0] * len(tiers)
                run_portions  = []
                for _r, run_sold in run_data:
                    rp        = safe_float(_r.iloc[ItemCol.COPIES_PRINTED])
                    run_start = cumul_printed + 1.0
                    run_end   = cumul_printed + max(rp, 0.0)
                    cumul_printed = run_end
                    portions = []
                    if rp > 0.0:
                        for i, (min_c, max_c, t_rate) in enumerate(tiers):
                            ov_s = max(run_start, float(min_c))
                            ov_e = min(run_end,   float(max_c))
                            if ov_e >= ov_s:
                                p = ov_e - ov_s + 1.0
                                portions.append((p, i, t_rate))
                                if run_sold > 0.0:
                                    tier_buckets[i] += run_sold * (p / rp)
                    run_portions.append(portions)
                tier_split = [
                    (bucket, tiers[i][2])
                    for i, bucket in enumerate(tier_buckets)
                    if bucket > 0.0
                ]
            else:
                run_portions = [[] for _ in run_data]
                tier_split   = []

            period_balance = None  # set at contract level after all ISBNs are processed

            # UNSOLD COPIES (new prints only)
            any_new_print = any(
                extract_year(r.iloc[ItemCol.DATE_PRINTED]) == report_year_from_period(period, self.year)
                for r, _ in run_data
            )
            unsold_copies = (
                max(0.0, total_copies_printed - total_copies_sold)
                if any_new_print and total_copies_printed > 0 else None
            )

            multiple_runs = len(group) > 1

            # ── Print run sub-rows (only when >1 reprint) ──────────────────────
            # Rules:
            #  • Run in a single tier with 0 sold  → print_run row (with retail/rate, advance on first)
            #  • Run in a single tier with sold > 0 → one print_run_tier row (with calc)
            #  • Run straddling ≥2 tiers            → one print_run_tier row per portion
            # Advance / balance always go to the very first row (first run, first tier).
            any_prt_row = False   # any print_run_tier row created
            if multiple_runs:

                for idx, (r, run_sold) in enumerate(run_data):
                    is_first_run = (idx == 0)
                    run_title    = safe_str(r.iloc[ItemCol.TITLE_EN]) or safe_str(r.iloc[ItemCol.TITLE_TH])
                    date_str     = format_date_printed(r.iloc[ItemCol.DATE_PRINTED])
                    run_printed  = safe_float(r.iloc[ItemCol.COPIES_PRINTED])
                    portions     = run_portions[idx]   # [(p_printed, tier_idx, rate), …]

                    # Title column rule: run 0 → EN, run 1 → TH, run 2+ → blank
                    if idx == 0:
                        run_title_col = title_display
                    elif idx == 1:
                        run_title_col = (strip_reprint_suffix(safe_str(r.iloc[ItemCol.TITLE_TH]))
                                         or title_th_sub)
                    else:
                        run_title_col = ''

                    # Single-tier zero-sold → context row with retail/rate; advance on first run
                    if not portions or (len(portions) == 1 and run_sold == 0.0):
                        want_adv = is_first_run and not adv_placed
                        if want_adv:
                            adv_placed = True
                        _run_rate    = (portions[0][2] if portions else fallback_rate) or 0.0
                        _run_amt_thb = run_sold * retail * _run_rate if run_sold > 0 else 0.0
                        if run_sold > 0:
                            any_prt_row = True
                        _run_amt_ccy = _run_amt_thb / amt_ex_rate if (amt_ex_rate and _run_amt_thb) else 0.0
                        rows.append({
                            'row_type':       'print_run',
                            'job':            safe_str(r.iloc[ItemCol.JOB]),  # every row
                            'isbn':           isbn,                            # every row
                            'title':          run_title_col,
                            'title_th':       '',  # no separate TH row; embedded in run 1
                            'copies_printed': run_printed,
                            'date_printed':   date_str,
                            'copies_sold':    run_sold,
                            'retail_price':   retail or None,
                            'royalty_rate':   _run_rate or None,
                            'amount_thb':     _run_amt_thb,
                            'amount_ccy':     _run_amt_ccy,
                            'currency':       currency,
                            'amt_currency':   amt_cur,
                            'amt_ex_rate':    amt_ex_rate,
                            'adv':            adv           if want_adv else None,
                            'adv_currency':   adv_cur,
                            'prev_balance':   prev_balance  if want_adv else None,
                            'balance_paid':   balance_paid  if want_adv else None,
                            'period_balance': period_balance if want_adv else None,
                            'unsold_copies':  unsold_copies  if want_adv else None,
                            'stock_account':  stock_account  if want_adv else None,
                            'status_2024':    pay_status if want_adv else None,
                            'is_ebook':       is_ebook,
                            'ex_rate':        ex_rate,
                        })
                        continue

                    # Straddling or has sold: one print_run_tier row per tier portion
                    any_prt_row = True
                    for p_idx, (p_printed, _t_idx, t_rate) in enumerate(portions):
                        p_sold  = run_sold * (p_printed / run_printed) if run_printed > 0.0 else 0.0
                        amt_thb = p_sold * retail * t_rate

                        # Advance always on first row (first run, first portion)
                        want_adv = (not adv_placed) and (p_idx == 0) and is_first_run
                        if want_adv:
                            adv_placed = True

                        rows.append({
                            'row_type':       'print_run_tier',
                            'job':            safe_str(r.iloc[ItemCol.JOB]) if p_idx == 0 else '',
                            'isbn':           isbn if p_idx == 0 else None,
                            'title':          run_title_col if p_idx == 0 else '',
                            'title_th':       '',  # no separate TH row; embedded in run 1
                            'copies_printed': p_printed,
                            'date_printed':   date_str,
                            'retail_price':   retail,
                            'royalty_rate':   t_rate,
                            'copies_sold':    p_sold,
                            'amount_thb':     amt_thb,
                            'amount_ccy':     amt_thb / amt_ex_rate if amt_ex_rate else 0.0,
                            'currency':       currency,
                            'amt_currency':   amt_cur,
                            'amt_ex_rate':    amt_ex_rate,
                            'adv':            adv           if want_adv else 0.0,
                            'adv_currency':   adv_cur,
                            'prev_balance':   prev_balance  if want_adv else 0.0,
                            'balance_paid':   balance_paid  if want_adv else 0.0,
                            'period_balance': period_balance if want_adv else None,
                            'unsold_copies':  unsold_copies  if want_adv else None,
                            'stock_account':  stock_account  if want_adv else None,
                            'status_2024':    pay_status if want_adv else None,
                            'is_ebook':       is_ebook,
                            'ex_rate':        ex_rate,
                        })

            # ── Tier rows ────────────────────────────────────────────────────────
            # Skip when print_run_tier rows already carry all calculation amounts.
            job_first       = safe_str(first_r.iloc[ItemCol.JOB])
            has_inline_calc = any_prt_row

            def make_tier_row(tc, tr, is_first_tier, want_adv=False):
                amt_thb = eb_net if is_ebook else tc * retail * tr
                return {
                    'row_type':       'tier',
                    'job':            '' if multiple_runs else (job_first if is_first_tier else ''),
                    'isbn':           None if (multiple_runs or is_ebook) else (isbn if is_first_tier else None),
                    'title':          '' if multiple_runs else (title_display if is_first_tier else ''),
                    'title_th':       '' if multiple_runs else (title_th_sub  if is_first_tier else ''),
                    'copies_printed': None if (multiple_runs or is_ebook) else (total_copies_printed if is_first_tier else None),
                    'date_printed':   (
                        '' if multiple_runs else
                        (format_date_printed(first_r.iloc[ItemCol.DATE_PRINTED]) if is_first_tier else '')
                    ),
                    'copies_sold':    None if is_ebook else tc,
                    'retail_price':   None if is_ebook else retail,
                    'royalty_rate':   tr,
                    'amount_thb':     amt_thb,
                    'amount_ccy':     amt_thb / amt_ex_rate if amt_ex_rate else 0.0,
                    'currency':       currency,
                    'amt_currency':   amt_cur,
                    'amt_ex_rate':    amt_ex_rate,
                    'adv':            adv           if want_adv else 0.0,
                    'adv_currency':   adv_cur,
                    'prev_balance':   prev_balance  if want_adv else 0.0,
                    'balance_paid':   balance_paid  if want_adv else 0.0,
                    'period_balance': period_balance if want_adv else None,
                    'unsold_copies':  unsold_copies  if want_adv else None,
                    'stock_account':  stock_account  if want_adv else None,
                    'status_2024':    pay_status if want_adv else None,
                    'is_ebook':       is_ebook,
                    'ex_rate':        ex_rate,
                }

            if not has_inline_calc:
                if tier_split:
                    for j, (tc, tr) in enumerate(tier_split):
                        _want_adv = (j == 0) and not adv_placed
                        if _want_adv:
                            adv_placed = True
                        rows.append(make_tier_row(tc, tr, is_first_tier=(j == 0), want_adv=_want_adv))
                else:
                    _want_adv = not adv_placed
                    if _want_adv:
                        adv_placed = True
                    rows.append(make_tier_row(total_copies_sold, fallback_rate, is_first_tier=True, want_adv=_want_adv))

            # Blank separator row between ISBN groups (matches sample layout)
            rows.append({'row_type': 'blank'})

        # Contract-level period balance = Previous Balance − Σ AMOUNT(THB) + Balance Paid
        # Placed once on the first non-blank row; covers all ISBNs in this contract.
        total_amount_thb = sum(
            (row.get('amount_thb') or 0.0)
            for row in rows
            if row.get('row_type') != 'blank'
        )
        first_data = next((r for r in rows if r.get('row_type') != 'blank'), None)
        if first_data is not None:
            raw_prev    = first_data.get('prev_balance') or 0.0
            raw_paid    = first_data.get('balance_paid') or 0.0
            contract_pb = (raw_prev - total_amount_thb + raw_paid) if (raw_prev != 0 or total_amount_thb > 0) else None
            pb_placed = False
            for row in rows:
                if row.get('row_type') == 'blank':
                    continue
                row['period_balance'] = contract_pb if not pb_placed else None
                pb_placed = True

        # Remove trailing blank row
        while rows and rows[-1].get('row_type') == 'blank':
            rows.pop()

        return rows

    # ── Main generate ─────────────────────────────────────────────────────────

    def generate_report(self, country, agency, publisher, period='annual', output_dir=None):
        if output_dir is None:
            output_dir = tempfile.mkdtemp()

        self._ensure_rates()

        agent_items = self.item[
            self._clean(self.item.iloc[:, ItemCol.AGENCY]) == agency
        ].copy()

        type_col = self._clean(agent_items.iloc[:, ItemCol.ANNUAL_BI])
        bi_mask  = type_col.str.upper().str.contains('BI')

        bi_items = agent_items[bi_mask]
        an_items = agent_items[~bi_mask]

        bi1_rows = self._build_rows(bi_items, 'bi1')    if period in ('all', 'bi1')    else []
        bi2_rows = self._build_rows(bi_items, 'bi2')    if period in ('all', 'bi2')    else []
        an_rows  = self._build_rows(an_items, 'annual') if period in ('all', 'annual') else []

        safe_agent = re.sub(r'[\\/:*?"<>|]', '_', agency)
        filename   = f"SalesReport_{safe_agent}_{period}.xlsx"
        out_path   = os.path.join(output_dir, filename)

        self._write_excel(out_path, country, agency, publisher, None, None, period,
                          bi1_rows, bi2_rows, an_rows)
        return out_path

    def generate_all_publishers(self, country, agency, period='annual', output_dir=None):
        if output_dir is None:
            output_dir = tempfile.mkdtemp()
        xlsx_dir = os.path.join(output_dir, 'xlsx')
        os.makedirs(xlsx_dir, exist_ok=True)

        for pub in self.get_publishers(agency, country):
            self.generate_report(country, agency, pub, period, xlsx_dir)

        zip_path = os.path.join(output_dir, f'SalesReports_{period}.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for fname in os.listdir(xlsx_dir):
                zf.write(os.path.join(xlsx_dir, fname), fname)
        return zip_path

    def _get_info_for_isbn(self, isbn: str):
        """Return (country, publisher, intermediate_agent, contract_expiry) for an ISBN."""
        self._init_intra_cols()
        if not isbn:
            return ('', '', '', None)
        for isbn_col in self._intra_isbn_cols:
            mask = self._clean(self.intra.iloc[:, isbn_col]).str.startswith(isbn)
            matched = self.intra[mask]
            if not matched.empty:
                r = matched.iloc[0]
                sell_off = r.iloc[IntraCol.SELL_OFF]
                exp_date = r.iloc[IntraCol.EXP_DATE]
                def _is_valid_date(v):
                    if v is None:
                        return False
                    s = safe_str(v).strip().lower()
                    return s and s not in ('nan', 'nat', '')
                contract_expiry = sell_off if _is_valid_date(sell_off) else (exp_date if _is_valid_date(exp_date) else None)
                return (
                    safe_str(r.iloc[IntraCol.COUNTRY]),
                    safe_str(r.iloc[IntraCol.PUBLISHER]),
                    safe_str(r.iloc[IntraCol.AGENT]),
                    contract_expiry,
                )
        return ('', '', '', None)

    def generate_all(self, period='annual', output_dir=None, isbn_filter=None, agency_filter=None):
        """Generate one Excel per intra contract (may cover multiple ISBNs), bundled as ZIP."""
        if output_dir is None:
            output_dir = tempfile.mkdtemp()

        self._ensure_rates()
        self._init_intra_cols()

        # ── Build contract map: intra row → list of ISBNs ────────────────────
        def _is_valid_date(v):
            s = safe_str(v).strip().lower()
            return s and s not in ('nan', 'nat', '')

        contracts = []          # list of dicts per intra row
        isbn_to_contract = {}   # isbn → index into contracts

        for _, intra_row in self.intra.iterrows():
            isbns = []
            for isbn_col in self._intra_isbn_cols:
                val = safe_str(intra_row.iloc[isbn_col]).strip()
                m = PRODUCT_CODE_RE.match(val)
                if m:
                    isbn_val = m.group(1)
                    if isbn_val not in isbns:
                        isbns.append(isbn_val)
            if not isbns:
                continue
            sell_off = intra_row.iloc[IntraCol.SELL_OFF]
            exp_date = intra_row.iloc[IntraCol.EXP_DATE]
            contract_expiry = sell_off if _is_valid_date(sell_off) else (exp_date if _is_valid_date(exp_date) else None)
            idx = len(contracts)
            contracts.append({
                'isbns':              isbns,
                'country':            safe_str(intra_row.iloc[IntraCol.COUNTRY]),
                'publisher':          safe_str(intra_row.iloc[IntraCol.PUBLISHER]),
                'intermediate_agent': safe_str(intra_row.iloc[IntraCol.AGENT]),
                'contract_expiry':    contract_expiry,
            })
            agent_val = safe_str(intra_row.iloc[IntraCol.AGENT])
            pub_val   = safe_str(intra_row.iloc[IntraCol.PUBLISHER])
            for isbn in isbns:
                if isbn not in isbn_to_contract:
                    isbn_to_contract[isbn] = idx
                else:
                    # ถ้า contract เดิม agent ว่าง แต่ row นี้มี agent → อัปเดต
                    existing_idx = isbn_to_contract[isbn]
                    if not contracts[existing_idx]['intermediate_agent'] and agent_val:
                        contracts[existing_idx]['intermediate_agent'] = agent_val
                        if pub_val:
                            contracts[existing_idx]['publisher'] = pub_val
                        if not contracts[existing_idx]['country']:
                            contracts[existing_idx]['country'] = safe_str(intra_row.iloc[IntraCol.COUNTRY])

        # ── ISBN filter ───────────────────────────────────────────────────────
        isbn_set = set(isbn_filter) if isbn_filter else None

        # ── Iterate item file by agency ───────────────────────────────────────
        agency_series = self._clean(self.item.iloc[:, ItemCol.AGENCY])
        all_agencies  = sorted(agency_series.unique())
        if agency_filter:
            all_agencies = [a for a in all_agencies if a == agency_filter]

        # Write xlsx to an isolated temp dir so old period files don't leak into this ZIP
        xlsx_tmp = tempfile.mkdtemp()

        for agency_raw in all_agencies:
            agent_items = self.item[agency_series == agency_raw].copy()
            if agent_items.empty:
                continue

            agency_label = agency_raw if len(agency_raw) > 3 else 'Direct Publisher'
            safe_agency  = re.sub(r'[\\/:*?"<>|]', '_', agency_label)[:40]
            agency_dir   = os.path.join(xlsx_tmp, safe_agency)

            isbn_series = self._clean(agent_items.iloc[:, ItemCol.ISBN])
            processed_contracts = set()
            orphan_isbns = []

            for isbn in sorted(isbn_series.unique()):
                if not isbn or not PRODUCT_CODE_RE.match(isbn):
                    continue
                if isbn_set and isbn not in isbn_set:
                    continue
                contract_idx = isbn_to_contract.get(isbn)
                if contract_idx is None:
                    orphan_isbns.append(isbn)
                    continue
                if contract_idx in processed_contracts:
                    continue
                processed_contracts.add(contract_idx)

                c = contracts[contract_idx]
                all_isbns  = c['isbns']
                # Collect item rows for ALL ISBNs in this contract (current agency)
                contract_items = agent_items[isbn_series.isin(all_isbns)].copy()

                safe_pub = re.sub(r'[\\/:*?"<>|]', '_', c['publisher'] or 'Unknown Publisher')[:40]
                pub_dir  = os.path.join(agency_dir, safe_pub)
                os.makedirs(pub_dir, exist_ok=True)

                type_col = self._clean(contract_items.iloc[:, ItemCol.ANNUAL_BI])
                bi_mask  = type_col.str.upper().str.contains('BI')

                bi1_rows = self._build_rows(contract_items[bi_mask],  'bi1')    if period in ('all', 'bi1')    else []
                bi2_rows = self._build_rows(contract_items[bi_mask],  'bi2')    if period in ('all', 'bi2')    else []
                an_rows  = self._build_rows(contract_items[~bi_mask], 'annual') if period in ('all', 'annual') else []

                if not any([bi1_rows, bi2_rows, an_rows]):
                    continue

                # File name: ISBN - EN title - TH title
                primary_items = contract_items[isbn_series.isin([all_isbns[0]])]
                if primary_items.empty:
                    primary_items = contract_items.iloc[:1]
                _row0       = primary_items.iloc[0]
                _th_raw     = strip_reprint_suffix(safe_str(_row0.iloc[ItemCol.TITLE_TH]))
                _th_clean   = re.sub(r'\s+\d+\.\d+', '', _th_raw).strip()
                _en_raw     = strip_reprint_suffix(safe_str(_row0.iloc[ItemCol.TITLE_EN]))
                # exclude values that look like agency names (end with ";")
                _en_clean   = '' if _en_raw.endswith(';') else _en_raw
                _label      = f"{_en_clean} - {_th_clean}" if _en_clean else _th_clean
                safe_name   = re.sub(r'[\\/:*?"<>|]', '_', f"{all_isbns[0]} - {_label}")[:100]

                self._write_excel(
                    os.path.join(pub_dir, f"{safe_name}.xlsx"),
                    c['country'], agency_label, c['publisher'],
                    c['intermediate_agent'], c['contract_expiry'], period,
                    bi1_rows, bi2_rows, an_rows,
                )

            # Orphaned ISBNs (not in any intra contract)
            for isbn in orphan_isbns:
                if isbn_set and isbn not in isbn_set:
                    continue
                isbn_items = agent_items[isbn_series == isbn].copy()
                country, publisher, intermediate_agent, contract_expiry = self._get_info_for_isbn(isbn)

                safe_pub = re.sub(r'[\\/:*?"<>|]', '_', publisher or 'Unknown Publisher')
                pub_dir  = os.path.join(agency_dir, safe_pub)
                os.makedirs(pub_dir, exist_ok=True)

                type_col = self._clean(isbn_items.iloc[:, ItemCol.ANNUAL_BI])
                bi_mask  = type_col.str.upper().str.contains('BI')

                bi1_rows = self._build_rows(isbn_items[bi_mask],  'bi1')    if period in ('all', 'bi1')    else []
                bi2_rows = self._build_rows(isbn_items[bi_mask],  'bi2')    if period in ('all', 'bi2')    else []
                an_rows  = self._build_rows(isbn_items[~bi_mask], 'annual') if period in ('all', 'annual') else []

                if not any([bi1_rows, bi2_rows, an_rows]):
                    continue

                _row0     = isbn_items.iloc[0]
                _th_clean = strip_reprint_suffix(safe_str(_row0.iloc[ItemCol.TITLE_TH]))
                _en_raw   = strip_reprint_suffix(safe_str(_row0.iloc[ItemCol.TITLE_EN]))
                _en_clean = '' if _en_raw.endswith(';') else _en_raw
                _label    = f"{_en_clean} - {_th_clean}" if _en_clean else _th_clean
                safe_name = re.sub(r'[\\/:*?"<>|]', '_', f"{isbn} - {_label}")[:100]
                self._write_excel(
                    os.path.join(pub_dir, f"{safe_name}.xlsx"),
                    country, agency_label, publisher, intermediate_agent, contract_expiry, period,
                    bi1_rows, bi2_rows, an_rows,
                )

        zip_path = os.path.join(output_dir, f'SalesReports_All_{period}.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for root, _dirs, files in os.walk(xlsx_tmp):
                for fname in files:
                    if fname.endswith('.xlsx'):
                        fpath   = os.path.join(root, fname)
                        arcname = os.path.relpath(fpath, xlsx_tmp)
                        zf.write(fpath, arcname)
        shutil.rmtree(xlsx_tmp, ignore_errors=True)

        with zipfile.ZipFile(zip_path, 'r') as zf:
            if not zf.namelist():
                os.remove(zip_path)
                period_label = {'bi1': 'BI-Annual 1', 'bi2': 'BI-Annual 2', 'annual': 'Annual'}.get(period, period)
                raise ValueError(f"ไม่พบข้อมูล {period_label} — ตรวจสอบว่าคอลัมน์ Annual/BI-Annual มีค่า 'BI' หรือไม่")

        return zip_path

    # ── Dashboard data ───────────────────────────────────────────────────────────

    def get_dashboard_data(self):
        self._ensure_rates()
        self._init_intra_cols()
        item = self.item

        # ── ISBN-level aggregation ────────────────────────────────────────────
        isbn_stats = {}
        for _, row in item.iterrows():
            isbn = safe_str(row.iloc[ItemCol.ISBN])
            if not isbn or not any(c.isdigit() for c in isbn):
                continue
            title = (strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_EN]))
                     or strip_reprint_suffix(safe_str(row.iloc[ItemCol.TITLE_TH])))
            agency = safe_str(row.iloc[ItemCol.AGENCY])
            if not agency or len(agency) <= 3:
                agency = 'Direct Publisher'

            paidtype   = safe_str(row.iloc[ItemCol.ANNUAL_BI]).upper()
            is_bi      = 'BI' in paidtype
            if is_bi:
                copies_sold = (safe_float(row.iloc[ItemCol.BI_H1_AMARIN])
                             + safe_float(row.iloc[ItemCol.BI_H1_ABOOK])
                             + safe_float(row.iloc[ItemCol.BI_H2_AMARIN])
                             + safe_float(row.iloc[ItemCol.BI_H2_ABOOK]))
            else:
                copies_sold = safe_float(row.iloc[ItemCol.ANNUAL_SOLD])

            copies_printed = safe_float(row.iloc[ItemCol.COPIES_PRINTED])
            retail         = safe_float(row.iloc[ItemCol.PRICE])
            rate           = normalize_rate(row.iloc[ItemCol.ROYALTY_RATE]) or self.flat_rate_for_isbn(isbn)
            royalty_thb    = copies_sold * retail * rate
            status_2024    = safe_str(row.iloc[ItemCol.STATUS_2024])
            prev_balance   = safe_float(row.iloc[ItemCol.PREV_BALANCE])

            if isbn not in isbn_stats:
                isbn_stats[isbn] = {
                    'title':          title,
                    'agency':         agency,
                    'copies_sold':    0.0,
                    'copies_printed': 0.0,
                    'royalty_thb':    0.0,
                    'status_2024':    status_2024,
                    'prev_balance':   prev_balance,
                }
            isbn_stats[isbn]['copies_sold']    += copies_sold
            isbn_stats[isbn]['copies_printed'] += copies_printed
            isbn_stats[isbn]['royalty_thb']    += royalty_thb

        total_isbns   = len(isbn_stats)
        with_sales    = sum(1 for v in isbn_stats.values() if v['copies_sold'] > 0)
        zero_sales    = sum(1 for v in isbn_stats.values() if v['copies_sold'] == 0)
        total_royalty = sum(v['royalty_thb'] for v in isbn_stats.values())

        # ── Advance status (per unique ISBN) ──────────────────────────────────
        advance_status = {'จ่ายแล้ว': 0, 'ค้างจ่าย': 0, 'ยังไม่เกิน ADV': 0}
        seen = set()
        for _, row in item.iterrows():
            isbn = safe_str(row.iloc[ItemCol.ISBN])
            if not isbn or isbn in seen:
                continue
            seen.add(isbn)
            s = safe_str(row.iloc[ItemCol.STATUS_2024])
            if s in advance_status:
                advance_status[s] += 1

        # ── Sell-off expiring (ปีของ report) ──────────────────────────────────
        expiring = set()
        for _, row in self.intra.iterrows():
            try:
                sell_yr = extract_year(row.iloc[IntraCol.SELL_OFF])
            except Exception:
                continue
            if sell_yr == self.year:
                for isbn_col in self._intra_isbn_cols:
                    v = safe_str(row.iloc[isbn_col])
                    if v:
                        expiring.add(v)
                        break

        # ── Agent aggregation ─────────────────────────────────────────────────
        agent_agg = {}
        for isbn, stats in isbn_stats.items():
            ag = stats['agency']
            if ag not in agent_agg:
                agent_agg[ag] = {'printed': 0.0, 'sold': 0.0, 'isbns': set(), 'royalty': 0.0}
            agent_agg[ag]['printed'] += stats['copies_printed']
            agent_agg[ag]['sold']    += stats['copies_sold']
            agent_agg[ag]['isbns'].add(isbn)
            agent_agg[ag]['royalty'] += stats['royalty_thb']

        agents = sorted([
            {
                'name':             ag,
                'isbn_count':       len(v['isbns']),
                'sell_through_pct': round(v['sold'] / v['printed'] * 100, 1) if v['printed'] > 0 else 0.0,
                'royalty_thb':      round(v['royalty'], 0),
            }
            for ag, v in agent_agg.items()
        ], key=lambda x: -x['royalty_thb'])

        # ── Top books & zero-sales books ──────────────────────────────────────
        all_books = sorted([
            {
                'isbn':             isbn,
                'title':            v['title'],
                'agency':           v['agency'],
                'copies_sold':      int(v['copies_sold']),
                'royalty_thb':      round(v['royalty_thb'], 0),
                'sell_through_pct': round(v['copies_sold'] / v['copies_printed'] * 100, 1)
                                    if v['copies_printed'] > 0 else 0.0,
                'status':           v['status_2024'],
            }
            for isbn, v in isbn_stats.items()
        ], key=lambda x: -x['royalty_thb'])

        top_books  = all_books[:10]
        zero_books = [b for b in all_books if b['copies_sold'] == 0][:10]

        # ── Totals ────────────────────────────────────────────────────────────
        agency_series    = self._clean(item.iloc[:, ItemCol.AGENCY])
        total_agencies   = len([a for a in agency_series.unique() if a and len(a) > 3])
        pub_series       = self._clean(self.intra.iloc[:, IntraCol.PUBLISHER])
        total_publishers = len([p for p in pub_series.unique()
                                 if p and p.lower() not in ('nan', '')])

        return {
            'summary': {
                'total_isbns':       total_isbns,
                'total_agencies':    total_agencies,
                'total_publishers':  total_publishers,
                'with_sales':        with_sales,
                'zero_sales':        zero_sales,
                'expiring_2025':     len(expiring),
                'report_year':       self.year,
                'total_royalty_thb': round(total_royalty, 0),
            },
            'advance_status':  advance_status,
            'agents':          agents,
            'top_books':       top_books,
            'zero_books':      zero_books,
            'all_books':       all_books,
        }

    # ── Excel writer ──────────────────────────────────────────────────────────

    def _write_excel(self, path, country, agency, publisher, intermediate_agent, contract_expiry, period,
                     bi1_rows, bi2_rows, an_rows):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sales Report'

        thin     = Side(style='thin')
        no_side  = Side(style=None)
        bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)
        ctr      = Alignment(horizontal='center', vertical='center', wrap_text=True)
        lft      = Alignment(horizontal='left',   vertical='center', wrap_text=True)
        rgt      = Alignment(horizontal='right',  vertical='center')
        NUM      = '#,##0.00'
        PCT      = '0.00%'
        TABLE_COLS = 18   # table ends at col R (จ่าย ค่าลิขสิทธิ์)

        def f(sz=10, bold=False, color='000000'):
            return Font(name='Cambria', size=sz, bold=bold, color=color)

        HDR_FILL = PatternFill('solid', fgColor='E0FFFF')
        HDR_FONT = Font(name='Cambria', size=10, bold=True, color='000000')
        SEC_FILL = PatternFill('solid', fgColor='D6E4F0')

        type_mark = {
            'bi1': '   ANNUAL               x  BI-ANNUAL',
            'bi2': '   ANNUAL               x  BI-ANNUAL',
        }.get(period, 'x  ANNUAL               BI-ANNUAL')

        period_end = period_end_label(period, self.year)

        # ── Rows 1-6: header block ────────────────────────────────────────────
        # Row 3: agency (the intermediate agent sending the report — e.g. Silkroad)
        # Row 4: licensor line — "SubAgent/Publisher" or just "Publisher" when
        #        intermediate_agent duplicates the agency (Silkroad is its own intra agent).
        def _effective_intermediate(ia, ag):
            if not ia:
                return ''
            # Normalise: strip trailing dots/commas, lowercase compare
            ia_clean = ia.rstrip('.,; ').lower()
            ag_clean = ag.rstrip('.,; ').lower()
            if ia_clean in ag_clean or ag_clean in ia_clean:
                return ''  # same entity — don't repeat
            return ia

        eff_ia = _effective_intermediate(intermediate_agent or '', agency)
        if eff_ia and publisher:
            licensor_line = f"{eff_ia}/{publisher}"
        else:
            licensor_line = eff_ia or publisher or ''

        row = 1
        for txt in ['SALES REPORT', 'Amarin Corporations PCL',
                    agency, licensor_line,
                    type_mark, period_end]:
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=TABLE_COLS)
            cell = ws.cell(row=row, column=3, value=txt)
            is_red = (row == 6)
            cell.font = Font(name='Cambria', size=16, bold=True, color='FF0000' if is_red else '000000')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row].height = 22
            row += 1

        # ── Rows 7-9: column headers ──────────────────────────────────────────
        all_rows_combined = bi1_rows + bi2_rows + an_rows
        raw_ccy     = next((r['currency']     for r in all_rows_combined if r.get('currency')),     'CCY')
        raw_amt_ccy = next((r['amt_currency'] for r in all_rows_combined if r.get('amt_currency')), raw_ccy)
        ccy_label     = 'JPY' if raw_ccy     == 'JYP' else raw_ccy
        amt_ccy_label = 'JPY' if raw_amt_ccy == 'JYP' else raw_amt_ccy

        # Exchange rate reference cell (T6) — used by col 10 formula =I{r}/$T$6
        _rate_val = next((r['amt_ex_rate'] for r in all_rows_combined if r.get('amt_ex_rate')), 1.0)
        _rate_cell = ws.cell(row=6, column=20, value=_rate_val)
        _rate_cell.number_format = '0.0000'
        _rate_cell.font = f(7, color='AAAAAA')
        RATE_REF = '$T$6'

        hdr_rows = [
            ['', '', 'TITLE', 'NO.OF',   'DATE',    'RETAIL',  'ROYALTY', 'NO.OF',
             'AMOUNT',  'AMOUNT',          'ADVANCED', 'PREVIOUS', 'BALANCE', 'PERIOD', 'NO.OF',  'Stock (เล่ม)', 'DIF',           'จ่าย'],
            ['', '', '',      'COPIES',  'PRINTED', 'PRICE',   'RATE',    'COPIES',
             '(THB)',   f'({amt_ccy_label})', 'PAYMENT',  'BALANCE',  'PAID',    'BALANCE', 'UNSOLD', 'คงเหลือ',       'Stock (เล่ม)',   'ค่าลิขสิทธิ์'],
            ['JOB', 'ISBN', '', 'PRINTED', '', '(THB)', '', 'SOLD',
             '', '', f'({ccy_label})', f'({ccy_label})', f'({ccy_label})', f'({ccy_label})', 'COPIES', 'Account',       'คงเหลือ Account', ''],
        ]
        hdr_start_row = row
        for hdr_idx, hdr in enumerate(hdr_rows):
            is_top    = hdr_idx == 0
            is_bottom = hdr_idx == 2
            for c_idx, val in enumerate(hdr, 1):
                cell = ws.cell(row=row, column=c_idx, value=val)
                cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = ctr
                cell.border = Border(
                    left=thin, right=thin,
                    top=thin if is_top else no_side,
                    bottom=thin if is_bottom else no_side,
                )
            ws.row_dimensions[row].height = 15
            row += 1

        # Merge A7:A9 (JOB) and B7:B9 (ISBN)
        for col, label in [(1, 'JOB'), (2, 'ISBN')]:
            ws.merge_cells(start_row=hdr_start_row, start_column=col,
                           end_row=hdr_start_row + 2, end_column=col)
            cell = ws.cell(row=hdr_start_row, column=col, value=label)
            cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = ctr
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # ── Section writer ────────────────────────────────────────────────────
        pb_row        = None   # Excel row where PERIOD BALANCE lives
        amount_rows   = []     # Excel rows that carry amount_thb (for SUM formula)

        def write_section(label, rows_data, sub_period):
            nonlocal row, pb_row, amount_rows
            if not rows_data:
                return

            total_thb = 0.0
            total_ccy = 0.0
            ex_rate_used = None
            ccy_used     = None

            for d in rows_data:
                def put(col, val, fmt=None, align=rgt):
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.font = f(9); cell.border = bdr; cell.alignment = align
                    if fmt:
                        cell.number_format = fmt

                if d['row_type'] == 'blank':
                    row += 1
                    continue

                # ── helpers shared by all row types ──────────────────────────
                r = row  # snapshot: row number for formulas
                use_formula = bool(d.get('copies_sold')) and not d.get('is_ebook')

                def _put_amounts():
                    if use_formula:
                        put(9,  f'=F{r}*G{r}*H{r}',      NUM)
                        put(10, f'=I{r}/{RATE_REF}',      NUM)
                        amount_rows.append(r)
                    else:
                        put(9,  d['amount_thb'] or None,  NUM)
                        put(10, d['amount_ccy'] or None,  NUM)
                        if d.get('amount_thb'):
                            amount_rows.append(r)

                def _track_pb():
                    nonlocal pb_row
                    if d.get('period_balance') is not None and pb_row is None:
                        pb_row = r
                    put(14, None, NUM)   # placeholder — filled after all sections

                if d['row_type'] == 'print_run':
                    put(1, d['job'],                    align=lft)
                    put(2, d['isbn'] or None,           align=lft)
                    put(3, d['title'],                  align=lft)
                    put(4, d['copies_printed'] or None, NUM)
                    put(5, d['date_printed'],           align=lft)
                    put(6, d['retail_price'] or None,   NUM)
                    put(7, d['royalty_rate'] or None,   PCT)
                    put(8, d['copies_sold'] or None,    NUM)
                    _put_amounts()
                    put(11, d['adv']          if d['adv'] is not None else None, NUM)
                    put(12, d['prev_balance'] if d['prev_balance'] is not None else None, NUM)
                    put(13, d['balance_paid'] if d['balance_paid'] is not None else None, NUM)
                    _track_pb()
                    put(15, d['unsold_copies'], NUM)
                    put(16, d['stock_account'] if d['stock_account'] is not None else None, NUM)
                    _stock = d.get('stock_account'); _unsold = d.get('unsold_copies')
                    put(17, (_stock - _unsold) if (_stock is not None and _unsold is not None) else None, NUM)
                    _st = d.get('status_2024')
                    put(18, (_st if _st else '-') if d['adv'] is not None else None, align=ctr)
                    total_thb   += d['amount_thb'] or 0.0
                    total_ccy   += d['amount_ccy'] or 0.0
                    ex_rate_used = d['ex_rate']
                    ccy_used     = d['currency']
                elif d['row_type'] == 'print_run_tier':
                    put(1, d['job'],          align=lft)
                    put(2, d['isbn'] or None, align=lft)
                    put(3, d['title'],        align=lft)
                    put(4, d['copies_printed'] or None,  NUM)
                    put(5, d['date_printed'],             align=lft)
                    put(6, d['retail_price']  or None,   NUM)
                    put(7, d['royalty_rate']  or None,   PCT)
                    put(8, d['copies_sold']   or None,   NUM)
                    _put_amounts()
                    put(11, d['adv']          or None,   NUM)
                    put(12, d['prev_balance'] or None,   NUM)
                    put(13, d['balance_paid'] or None,   NUM)
                    _track_pb()
                    put(15, d['unsold_copies'],           NUM)
                    put(16, d['stock_account'] or None,  NUM)
                    _stock = d.get('stock_account'); _unsold = d.get('unsold_copies')
                    put(17, (_stock - _unsold) if (_stock is not None and _unsold is not None) else None, NUM)
                    _st = d.get('status_2024')
                    put(18, (_st if _st else '-') if d.get('status_2024') is not None else None, align=ctr)
                    total_thb   += d['amount_thb']
                    total_ccy   += d['amount_ccy']
                    ex_rate_used = d['ex_rate']
                    ccy_used     = d['currency']
                else:
                    put(1, d['job'],          align=lft)
                    put(2, d['isbn'] or None, align=lft)
                    put(3,  d['title'],                        align=lft)
                    put(4,  d['copies_printed'] or None,       NUM)
                    put(5,  d['date_printed'],                 align=lft)
                    put(6,  d['retail_price']   or None,       NUM)
                    put(7,  d['royalty_rate']   or None,       PCT)
                    put(8,  d['copies_sold']    or None,       NUM)
                    _put_amounts()
                    put(11, d['adv']            or None,       NUM)
                    put(12, d['prev_balance']   or None,       NUM)
                    put(13, d['balance_paid']   or None,       NUM)
                    _track_pb()
                    put(15, d['unsold_copies'],                NUM)
                    put(16, d['stock_account']  or None,       NUM)
                    _stock = d.get('stock_account'); _unsold = d.get('unsold_copies')
                    put(17, (_stock - _unsold) if (_stock is not None and _unsold is not None) else None, NUM)
                    _st = d.get('status_2024')
                    put(18, (_st if _st else '-') if _st is not None else None, align=ctr)
                    total_thb   += d['amount_thb']
                    total_ccy   += d['amount_ccy']
                    ex_rate_used = d['ex_rate']
                    ccy_used     = d['currency']

                row += 1

                if d.get('title_th'):
                    put(3, d['title_th'], align=lft)
                    for c in range(1, 19):
                        if c != 3:
                            ws.cell(row=row, column=c).border = bdr
                    row += 1

            # "of net Receipt" note for e-books
            if any(d.get('is_ebook') for d in rows_data):
                ws.cell(row=row, column=7, value='of net Receipt').font = f(8)
                row += 1


            # Exchange rate note
            if ex_rate_used and ccy_used:
                ccy_lbl = 'JPY' if ccy_used == 'JYP' else ccy_used
                ws.cell(row=row, column=6,
                        value=f'*  at current exchange rate of {ccy_lbl} 1 per {ex_rate_used:.4f}  Baht').font = f(8)
                row += 1

            # Contract expiry note
            if contract_expiry is not None:
                try:
                    if hasattr(contract_expiry, 'day'):
                        expiry_str = f"{contract_expiry.day}/{contract_expiry.month}/{contract_expiry.year}"
                    else:
                        dt = pd.to_datetime(safe_str(contract_expiry))
                        expiry_str = f"{dt.day}/{dt.month}/{dt.year}" if pd.notna(dt) else safe_str(contract_expiry)
                    ws.cell(row=row, column=6,
                            value=f'* contract expires on {expiry_str}').font = f(8)
                    row += 1
                except Exception:
                    pass

            row += 1

        yr = self.year
        write_section(f'BI-ANNUAL {yr}.1  (January – June {yr})',      bi1_rows, 'bi1')
        write_section(f'BI-ANNUAL {yr}.2  (July – December {yr})',     bi2_rows, 'bi2')
        write_section(f'ANNUAL {yr}  (January – December {yr})',       an_rows,  'annual')

        # ── Col 14: PERIOD BALANCE formula ────────────────────────────────────
        # Written after all sections so the SUM covers every amount row.
        if pb_row is not None:
            cell14 = ws.cell(row=pb_row, column=14)
            cell14.number_format = '#,##0.00'
            cell14.font = Font(name='Cambria', size=10)
            cell14.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                   top=Side(style='thin'),  bottom=Side(style='thin'))
            cell14.alignment = Alignment(horizontal='right', vertical='center')
            if amount_rows:
                sum_parts = '+'.join(f'I{r}' for r in amount_rows)
                cell14.value = f'=L{pb_row}-({sum_parts})+M{pb_row}'
            else:
                cell14.value = f'=L{pb_row}+M{pb_row}'

        # ── Column widths ─────────────────────────────────────────────────────
        for i, w in enumerate(
            [12, 16, 42, 11, 13, 10, 9, 9, 13, 13, 13, 13, 13, 13, 10, 13, 13, 12], 1
        ):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A10'
        wb.save(path)
