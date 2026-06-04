"""
Legacy Report Converter
=======================
อ่านไฟล์ report เก่า (XLS/XLSX) จาก DataSale/Report XXXX/
แล้วสร้างไฟล์ Excel ใหม่ format v0.45 แยกตามปี/Agent/Publisher/หนังสือ

Output structure:
  {year}/{agent}/{publisher}/{title_en} - {title_th}.xlsx

รองรับ 4 formats:
  A – Japan XLS    : SALES REPORT ที่ col 0, 1 book/section
  B – Standard XLSX: SALES REPORT ที่ col 1, 1-N books/section
  C – TLL/China    : SALES REPORT ที่ col 0, N books/section (no agent in parens)
  D – E-book       : "EBOOK/CHAPTER SALES REPORT" (parse best-effort)
"""

import os
import re
import shutil
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Optional

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Helpers ───────────────────────────────────────────────────────────────────

def _sv(v) -> str:
    """safe string"""
    if v is None:
        return ''
    s = str(v).strip()
    return '' if s.lower() in ('none', 'nan', 'nat') else s


def _sf(v) -> Optional[float]:
    """safe float – returns None on failure or NaN"""
    try:
        f = float(v)
        return None if (f != f) else f
    except Exception:
        return None


def _read_rows(path: Path):
    """Read XLS or XLSX → list of list of values"""
    ext = path.suffix.lower()
    if ext == '.xls':
        if not HAS_XLRD:
            raise RuntimeError("ต้องติดตั้ง xlrd: pip install xlrd")
        wb = xlrd.open_workbook(str(path))
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(r, c) for c in range(ws.ncols)]
                for r in range(ws.nrows)]
    else:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        return [[ws.cell(r, c).value
                 for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]


_SECTION_HEADERS = {
    'SALES REPORT',
    'EBOOK/CHAPTER SALES REPORT',
    'E-BOOK SALES REPORT',
    'EBOOK SALES REPORT',
}


def _detect_offset(rows) -> int:
    """หา column offset ของ SALES REPORT (0 หรือ 1)"""
    for row in rows[:5]:
        for i, v in enumerate(row[:5]):
            if _sv(v).upper() in _SECTION_HEADERS:
                return i
    return 0

def _find_section_starts(rows, offset: int):
    """Return list of row indices ที่เป็นจุดเริ่ม section ใหม่"""
    return [
        i for i, row in enumerate(rows)
        if len(row) > offset
        and _sv(row[offset]).upper() in _SECTION_HEADERS
    ]


def _is_ebook_section(rows, start: int, offset: int) -> bool:
    """True ถ้า section เป็น E-Book format (ไม่มีคอลัมน์ copies_printed)"""
    header = _sv(rows[start][offset]).upper() if len(rows[start]) > offset else ''
    if 'E-BOOK' in header or 'EBOOK' in header:
        return True
    # ตรวจจาก column header row: ถ้าไม่มี "NO.OF COPIES PRINTED" แต่มี "PUBLICATION DATE" ตรงๆ
    for i in range(start + 1, min(start + 10, len(rows))):
        row_vals = [_sv(v).upper() for v in rows[i]]
        if 'PUBLICATION' in row_vals:
            return True
    return False


# ── Period parsing ─────────────────────────────────────────────────────────────

def _extract_year(text: str) -> Optional[int]:
    """ดึงปีจาก period_end string"""
    m = re.findall(r'\b(20\d{2})\b', text)
    if m:
        return int(m[-1])   # ใช้ปีท้ายสุด (end period)
    return None


def _extract_period_code(period_type: str, period_end: str) -> str:
    """
    'annual' | 'bi1' | 'bi2'
    BI-H1 = Jan-Jun, BI-H2 = Jul-Dec
    """
    pt = period_type.upper()
    pe = period_end.upper()

    # ถ้า period_end บอกชัดเจน
    for kw in ('JUNE', 'JUN', '30 JUNE', 'JUNE 30'):
        if kw in pe:
            return 'bi1'
    for kw in ('DECEMBER', 'DEC', '31 DECEMBER', 'DECEMBER 31'):
        if kw in pe:
            # BI-ANNUAL ending December = H2
            if 'BI' in pt:
                return 'bi2'
            return 'annual'

    # fallback จาก period_type
    if 'BI' in pt:
        # BI-ANNUAL ไม่รู้ H1/H2 → ใช้ปีช่วยตัดสิน
        return 'bi1'   # default H1
    return 'annual'


def _period_label(code: str, year: int) -> str:
    labels = {
        'annual': f'Annual {year}  (January – December {year})',
        'bi1':    f'BI-Annual {year}.1  (January – June {year})',
        'bi2':    f'BI-Annual {year}.2  (July – December {year})',
    }
    return labels.get(code, f'{code} {year}')


def _period_end_str(code: str, year: int) -> str:
    ends = {
        'annual': f'For the Period Ended December 31, {year}',
        'bi1':    f'For the Period Ended June 30, {year}',
        'bi2':    f'For the Period Ended December 31, {year}',
    }
    return ends.get(code, f'For the Period Ended December 31, {year}')


def _type_mark(code: str) -> str:
    if code in ('bi1', 'bi2'):
        return '   ANNUAL               x  BI-ANNUAL'
    return 'x  ANNUAL               BI-ANNUAL'


# ── Header extraction ─────────────────────────────────────────────────────────

def _extract_header(rows, start: int, end: int, offset: int):
    """
    สกัด publisher, agency, period_type, period_end จาก section header
    ทนทานต่อหลาย format โดยใช้ pattern matching
    """
    publisher = ''
    agency    = ''
    period_type = ''
    period_end  = ''

    for i in range(start + 1, min(start + 8, end)):
        if i >= len(rows):
            break
        row = rows[i]
        val = _sv(row[offset]) if len(row) > offset else ''
        if not val:
            continue

        val_u = val.upper()

        # period end line
        if ('FOR THE PERIOD' in val_u
                or re.search(r'\b(JANUARY|JUNE|JULY|DECEMBER)\b', val_u)
                and re.search(r'\b20\d{2}\b', val)):
            period_end = val
            continue

        # period type line
        if 'ANNUAL' in val_u and len(val) < 80:
            period_type = val
            continue

        # skip Amarin line
        if 'AMARIN' in val_u:
            continue

        # skip header labels
        if val_u in ('TITLE', 'NO.OF', 'COPIES', 'PRINTED',
                     'DATE', 'RETAIL', 'PRICE', 'ROYALTY', 'RATE',
                     'SOLD', 'AMOUNT', 'PAYMENT', 'BALANCE', 'PAID', 'UNSOLD'):
            continue

        # publisher / agency line
        if not publisher:
            m = re.match(r'^(.+?)\s*\((.+)\)\s*$', val)
            if m:
                publisher = m.group(1).strip()
                agency    = m.group(2).strip()
            else:
                publisher = val
        elif not agency:
            agency = val

    return publisher, agency, period_type, period_end


# ── Data row detection ────────────────────────────────────────────────────────

_HEADER_LABELS = {
    'TITLE', 'NO.OF', 'COPIES', 'PRINTED', 'DATE', 'RETAIL',
    'PRICE', 'ROYALTY', 'RATE', 'SOLD', 'AMOUNT', 'PAYMENT',
    'BALANCE', 'PAID', 'UNSOLD', 'BAHT', 'BAHT)', 'STOCK',
    'BI-ANNUAL', 'ANNUAL', 'SALES REPORT', 'EBOOK/CHAPTER SALES REPORT',
    'AMARIN', 'OF NET', 'RECEIPT', 'FOR THE',
}

def _is_data_row(row, offset: int) -> bool:
    """
    True ถ้า row มี title ที่ col offset และมีตัวเลข (copies_printed หรือ copies_sold)
    """
    if len(row) <= offset:
        return False
    title = _sv(row[offset])
    if not title:
        return False
    if title.upper() in _HEADER_LABELS:
        return False
    # ต้องมีตัวเลขอย่างน้อย 1 ใน cols ถัดไป
    for c in range(offset + 1, min(offset + 10, len(row))):
        if _sf(row[c]) is not None:
            return True
    return False


def _is_subtitle_row(row, offset: int) -> bool:
    """True ถ้า row มีแค่ title (ไม่มีตัวเลข) — เป็น TH/romanized title sub-row"""
    if len(row) <= offset:
        return False
    title = _sv(row[offset])
    if not title:
        return False
    if title.upper() in _HEADER_LABELS:
        return False
    for c in range(offset + 1, min(offset + 10, len(row))):
        if _sf(row[c]) is not None:
            return False
    return True


# ── Currency detection ────────────────────────────────────────────────────────

def _extract_currency(rows, start: int, section_end: int) -> str:
    for i in range(start, min(section_end, start + 12)):
        if i >= len(rows):
            break
        for v in rows[i]:
            m = re.search(r'\((US\$|JPY|YEN|EUR|GBP|CNY|KRW|TWD|USD)\)', _sv(v), re.I)
            if m:
                c = m.group(1).upper()
                return {'US$': 'USD', 'YEN': 'JPY'}.get(c, c)
    return 'USD'


def _extract_exchange_rate(rows, start: int, section_end: int) -> Optional[float]:
    for i in range(start, min(section_end, len(rows))):
        for v in rows[i]:
            s = _sv(v)
            m = re.search(r'exchange rate[^\d]*(\d+[\.,]\d+)', s, re.I)
            if m:
                return _sf(m.group(1).replace(',', '.'))
    return None


def _extract_expiry(rows, start: int, section_end: int) -> str:
    for i in range(start, min(section_end, len(rows))):
        for v in rows[i]:
            s = _sv(v)
            if 'contract expires' in s.lower() or re.match(r'exp\.?\s', s, re.I):
                m = re.search(r'(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})', s)
                if m:
                    return m.group(1)
    return ''


# ── Book data extraction ───────────────────────────────────────────────────────

def _parse_book_from_data_row(data_row, sub_rows, offset: int,
                               currency: str, adv_pool: dict,
                               is_ebook_layout: bool = False) -> dict:
    """
    สร้าง book dict จาก 1 data_row + sub_rows (TH / romanized titles)
    adv_pool: shared advance ถ้าเป็น contract ที่มีหลาย volume
    """
    o = offset
    d = data_row

    def get(col_offset, fallback=None):
        idx = o + col_offset
        return d[idx] if idx < len(d) else fallback

    title_en = _sv(get(0))

    if is_ebook_layout:
        # E-Book format: ไม่มีคอลัมน์ copies_printed และ amount_ccy
        # layout: title | pub_date | retail | rate | copies_sold | amount_thb | adv | prev_bal | bal_paid | period_bal
        copies_printed = None
        date_printed   = _sv(get(1))
        retail_price   = _sf(get(2))
        royalty_rate   = _sf(get(3))
        copies_sold    = _sf(get(4))
        amount_thb     = _sf(get(5))
        amount_ccy     = None
        adv            = _sf(get(6))
        prev_balance   = _sf(get(7))
        balance_paid   = _sf(get(8))
        period_balance = _sf(get(9))
        unsold         = None
    else:
        # Standard format: title | copies_printed | date | retail | rate | sold | amount_thb | amount_ccy | adv | prev_bal | bal_paid | ? | period_bal | unsold
        copies_printed = _sf(get(1))
        date_printed   = _sv(get(2))
        retail_price   = _sf(get(3))
        royalty_rate   = _sf(get(4))
        copies_sold    = _sf(get(5))
        amount_thb     = _sf(get(6))
        amount_ccy     = _sf(get(7))
        adv            = _sf(get(8))
        prev_balance   = _sf(get(9))
        balance_paid   = _sf(get(10))
        period_balance = _sf(get(12))
        unsold         = _sf(get(13))

    # format date
    if date_printed and re.match(r'\d{4}-\d{2}-\d{2}', date_printed):
        try:
            dt = datetime.strptime(date_printed[:10], '%Y-%m-%d')
            date_printed = dt.strftime('%b-%y')
        except Exception:
            pass

    # extract TH title from sub_rows (Thai script)
    title_th = ''
    for sub in sub_rows:
        t = _sv(sub[o]) if len(sub) > o else ''
        if re.search(r'[ก-๙]', t):
            title_th = t
            break

    # currency fix
    if not currency or currency == 'USD':
        # try royalty_rate string for hints like "25%\n* of net receipt"
        rr_str = _sv(get(4))
        if 'YEN' in rr_str.upper() or 'JPY' in rr_str.upper():
            currency = 'JPY'

    # normalize royalty rate
    if royalty_rate and royalty_rate > 1:
        royalty_rate = royalty_rate / 100.0

    # adv pool — ใน contract หลาย volume, adv/prev_balance อยู่แค่ใน row แรก
    if adv is not None and adv != 0:
        adv_pool['adv']          = adv
        adv_pool['prev_balance'] = prev_balance
        adv_pool['balance_paid'] = balance_paid
    elif adv_pool.get('used'):
        adv = None; prev_balance = None; balance_paid = None
    else:
        adv          = adv_pool.get('adv', adv)
        prev_balance = adv_pool.get('prev_balance', prev_balance)
        balance_paid = adv_pool.get('balance_paid', balance_paid)
    adv_pool['used'] = True

    return {
        'title_en':      title_en,
        'title_th':      title_th,
        'copies_printed': copies_printed,
        'date_printed':  date_printed,
        'retail_price':  retail_price,
        'royalty_rate':  royalty_rate,
        'copies_sold':   copies_sold,
        'amount_thb':    amount_thb,
        'amount_ccy':    amount_ccy,
        'currency':      currency,
        'adv':           adv,
        'prev_balance':  prev_balance,
        'balance_paid':  balance_paid,
        'period_balance': period_balance,
        'unsold':        unsold,
    }


# ── Main section parser ───────────────────────────────────────────────────────

def _parse_section(rows, start: int, section_end: int, offset: int,
                   agent_folder: str) -> list:
    """
    Parse 1 section → list of book dicts (อาจมีหลาย book ใน TLL format)
    Each dict: publisher, agency, year, period_code, period_end_str, ...book fields
    """
    is_ebook        = _is_ebook_section(rows, start, offset)
    is_ebook_layout = is_ebook  # E-Book sections มี column layout ต่างออกไป

    publisher, agency, period_type, period_end_raw = \
        _extract_header(rows, start, section_end, offset)

    if not agency:
        agency = agent_folder
    if not publisher:
        publisher = 'Unknown Publisher'

    year = _extract_year(period_end_raw) or 2024
    period_code = _extract_period_code(period_type, period_end_raw)
    currency    = _extract_currency(rows, start, section_end)
    ex_rate     = _extract_exchange_rate(rows, start, section_end)
    expiry      = _extract_expiry(rows, start, section_end)

    # ── find data rows ────────────────────────────────────────────────────────
    books = []
    adv_pool = {}

    i = start + 1
    while i < section_end:
        row = rows[i]
        if _is_data_row(row, offset):
            # collect sub-rows (subtitle, romanized title)
            sub_rows = []
            j = i + 1
            while j < section_end and j < i + 5:
                if _is_data_row(rows[j], offset):
                    break
                if _is_subtitle_row(rows[j], offset):
                    sub_rows.append(rows[j])
                j += 1

            book = _parse_book_from_data_row(row, sub_rows, offset,
                                              currency, adv_pool,
                                              is_ebook_layout=is_ebook_layout)
            book.update({
                'publisher':    publisher,
                'agency':       agency,
                'year':         year,
                'period_code':  period_code,
                'period_end_raw': period_end_raw,
                'ex_rate':      ex_rate,
                'contract_expiry': expiry,
                'is_ebook':     is_ebook,
            })
            books.append(book)
            i = j
        else:
            i += 1

    return books


# ── Parse full file ───────────────────────────────────────────────────────────

def parse_legacy_file(path: Path, agent_folder: str) -> list:
    """
    Parse ไฟล์ report เก่า 1 ไฟล์ → list ของ book dicts
    """
    try:
        rows   = _read_rows(path)
    except Exception as e:
        raise RuntimeError(f"อ่านไฟล์ล้มเหลว: {e}")

    offset   = _detect_offset(rows)
    sections = _find_section_starts(rows, offset)

    if not sections:
        return []

    all_books = []
    for idx, s_start in enumerate(sections):
        s_end = sections[idx + 1] if idx + 1 < len(sections) else len(rows)
        try:
            books = _parse_section(rows, s_start, s_end, offset, agent_folder)
            all_books.extend(books)
        except Exception:
            pass

    return all_books


# ── Excel writer (v0.45 format) ───────────────────────────────────────────────

def write_legacy_excel(path: str, book: dict):
    """
    เขียน 1 book → Excel file format v0.45
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    thin = Side(style='thin')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    ctr  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lft  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    rgt  = Alignment(horizontal='right',  vertical='center')
    NUM  = '#,##0.00'
    PCT  = '0.00%'

    def f(sz=9, bold=False, color='000000'):
        return Font(name='Arial', size=sz, bold=bold, color=color)

    HDR_FILL = PatternFill('solid', fgColor='1F4E79')
    HDR_FONT = Font(name='Arial', size=9, bold=True, color='FFFFFF')

    year        = book.get('year', 2024)
    period_code = book.get('period_code', 'annual')
    publisher   = book.get('publisher', '')
    agency      = book.get('agency', '')

    # ── Rows 1-6: header block ────────────────────────────────────────────────
    header_lines = [
        'SALES REPORT',
        'Amarin Corporations PCL',
        agency,
        publisher,
        _type_mark(period_code),
        _period_end_str(period_code, year),
    ]
    for i, txt in enumerate(header_lines, 1):
        ws.cell(row=i, column=3, value=txt).font = f(10, bold=(i <= 2))

    # exchange rate reference cell T6
    ex_rate  = book.get('ex_rate') or 1.0
    currency = book.get('currency', 'USD')
    ccy_label = 'JPY' if currency == 'JYP' else currency

    rate_cell = ws.cell(row=6, column=20, value=ex_rate)
    rate_cell.number_format = '0.0000'
    rate_cell.font = f(7, color='AAAAAA')

    # ── Rows 7-9: column headers ──────────────────────────────────────────────
    hdr_rows = [
        ['', '', 'TITLE', 'NO.OF',   'DATE',    'RETAIL',  'ROYALTY', 'NO.OF',
         'AMOUNT',    'AMOUNT',        'ADVANCED', 'PREVIOUS', 'BALANCE', 'PERIOD',
         'NO.OF',  'Stock (เล่ม)', 'DIF',  'จ่าย'],
        ['', '', '',      'COPIES',  'PRINTED', 'PRICE',   'RATE',    'COPIES',
         '(THB)',     f'({ccy_label})', 'PAYMENT',  'BALANCE',  'PAID',    'BALANCE',
         'UNSOLD', 'คงเหลือ',    'Stock (เล่ม)', 'ค่าลิขสิทธิ์'],
        ['JOB', 'ISBN', '', 'PRINTED', '', '(THB)', '', 'SOLD',
         '', '', f'({ccy_label})', f'({ccy_label})', f'({ccy_label})', f'({ccy_label})',
         'COPIES', 'Account', 'คงเหลือ Account', ''],
    ]
    row_num = 7
    for hdr in hdr_rows:
        for c_idx, val in enumerate(hdr, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = ctr; cell.border = bdr
        row_num += 1

    # ── Data row ──────────────────────────────────────────────────────────────
    def put(col, val, fmt=None, align=rgt):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = f(9); cell.border = bdr; cell.alignment = align
        if fmt:
            cell.number_format = fmt

    title_en = book.get('title_en', '')
    title_th = book.get('title_th', '')

    royalty_rate   = book.get('royalty_rate')
    copies_sold    = book.get('copies_sold')
    retail_price   = book.get('retail_price')
    amount_thb     = book.get('amount_thb')
    amount_ccy     = book.get('amount_ccy')
    adv            = book.get('adv')
    prev_balance   = book.get('prev_balance')
    balance_paid   = book.get('balance_paid')
    period_balance = book.get('period_balance')
    copies_printed = book.get('copies_printed')
    date_printed   = book.get('date_printed', '')
    unsold         = book.get('unsold')

    put(1,  '',            align=lft)   # JOB
    put(2,  '',            align=lft)   # ISBN
    put(3,  title_en,      align=lft)
    put(4,  copies_printed, NUM)
    put(5,  date_printed,  align=lft)
    put(6,  retail_price,  NUM)
    put(7,  royalty_rate,  PCT)
    put(8,  copies_sold,   NUM)
    put(9,  amount_thb,    NUM)
    put(10, amount_ccy,    NUM)
    put(11, adv,           NUM)
    put(12, prev_balance,  NUM)
    put(13, balance_paid,  NUM)
    put(14, period_balance, NUM)
    put(15, unsold,        NUM)
    put(16, None,          NUM)   # Stock Account
    put(17, None,          NUM)   # DIF
    put(18, None,          align=ctr)
    data_row_num = row_num
    row_num += 1

    # TH title sub-row
    if title_th:
        put(3, title_th, align=lft)
        for c in range(1, 19):
            if c != 3:
                ws.cell(row=row_num, column=c).border = bdr
        row_num += 1

    # exchange rate note
    if ex_rate and currency and ex_rate != 1.0:
        ws.cell(row=row_num, column=6,
                value=f'*  at current exchange rate of {ccy_label} 1 per {ex_rate:.4f}  Baht'
                ).font = f(8)
        row_num += 1

    # contract expiry note
    expiry = book.get('contract_expiry', '')
    if expiry:
        ws.cell(row=row_num, column=6,
                value=f'* contract expires on {expiry}').font = f(8)
        row_num += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    for i, w in enumerate(
        [12, 16, 42, 11, 13, 10, 9, 9, 13, 13, 13, 13, 13, 13, 10, 13, 13, 12], 1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A10'
    wb.save(path)


# ── Safe filename helpers ─────────────────────────────────────────────────────

def _safe(text: str, max_len: int = 60) -> str:
    return re.sub(r'[\\/:*?"<>|]', '_', text).strip()[:max_len]


# ── Folder converter ──────────────────────────────────────────────────────────

def convert_datasale_folder(source_dir: str, output_dir: str,
                             progress_cb=None) -> tuple:
    """
    อ่านไฟล์ทั้งหมดจาก source_dir → เขียน Excel ใหม่ใน output_dir
    Returns: (zip_path, stats_dict)

    progress_cb(current, total, message) — optional callback สำหรับ progress
    """
    source = Path(source_dir)
    xlsx_tmp = tempfile.mkdtemp()

    stats = {'total_files': 0, 'total_books': 0, 'errors': []}

    all_files = [f for f in sorted(source.rglob('*.xls*'))
                 if not f.name.startswith('~')]
    total_files = len(all_files)

    if progress_cb:
        progress_cb(0, total_files, f'พบไฟล์ทั้งหมด {total_files} ไฟล์ กำลังเริ่มต้น...')

    # walk: region/agent/file.xls*
    for file_idx, fpath in enumerate(all_files):
        stats['total_files'] += 1

        # agent name = ชื่อ folder พ่อ (หรือ 2 levels ขึ้นไป)
        agent_folder = fpath.parent.name

        if progress_cb:
            progress_cb(file_idx + 1, total_files,
                        f'({file_idx+1}/{total_files}) {fpath.name}',
                        books_out=stats['total_books'])

        try:
            books = parse_legacy_file(fpath, agent_folder)
        except Exception as e:
            stats['errors'].append((fpath.name, str(e)))
            continue

        for book in books:
            if not book.get('title_en') and not book.get('title_th'):
                continue

            year        = book.get('year', 2024)
            period_code = book.get('period_code', 'annual')
            agency_name = book.get('agency', agent_folder) or agent_folder
            publisher   = book.get('publisher', 'Unknown')

            title_en = _safe(book.get('title_en', ''), 50)
            title_th = _safe(book.get('title_th', ''), 30)
            title_label = f"{title_en} - {title_th}" if title_en and title_th else (title_en or title_th)
            title_label = title_label[:70]

            # output path: {xlsx_tmp}/{year}/{agent}/{publisher}/{title}_{period}.xlsx
            out_dir = Path(xlsx_tmp) / str(year) / _safe(agency_name, 40) / _safe(publisher, 40)
            out_dir.mkdir(parents=True, exist_ok=True)

            period_suffix = {'annual': 'Annual', 'bi1': 'BI-H1', 'bi2': 'BI-H2'}.get(period_code, period_code)
            base_fname = f"{_safe(title_label, 80)}_{period_suffix}"
            # ป้องกัน filename ชนกัน — เพิ่ม _2, _3 ... ถ้าซ้ำ
            fname = f"{base_fname}.xlsx"
            counter = 2
            while (out_dir / fname).exists():
                fname = f"{base_fname}_{counter}.xlsx"
                counter += 1
            out_path = str(out_dir / fname)

            try:
                write_legacy_excel(out_path, book)
                stats['total_books'] += 1
            except Exception as e:
                stats['errors'].append((fpath.name + ' > ' + title_label, str(e)))

    # zip
    if progress_cb:
        progress_cb(total_files, total_files,
                    f'กำลัง zip {stats["total_books"]} ไฟล์...')

    zip_path = os.path.join(output_dir, 'legacy_reports.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(xlsx_tmp):
            for fname in files:
                if fname.endswith('.xlsx'):
                    fp = os.path.join(root, fname)
                    arcname = os.path.relpath(fp, xlsx_tmp)
                    zf.write(fp, arcname)

    shutil.rmtree(xlsx_tmp, ignore_errors=True)
    return zip_path, stats
