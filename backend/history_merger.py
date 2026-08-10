"""
History Merger
==============
รวม report Excel ปีปัจจุบัน (จาก generate_all) กับไฟล์ย้อนหลัง (จาก legacy_converter)
เป็นไฟล์เดียวที่มีทุกปีเรียงจากเก่าไปใหม่

Matching key: title_en (normalized lowercase)
  - report filename:  {ISBN} - {title_en} - {title_th}.xlsx
  - legacy filename:  {title_en} - {title_th}_{period}.xlsx  or  {title_en}_{period}.xlsx
"""
import re
import shutil
import tempfile
import zipfile
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl


# ── Title extraction helpers ────────────────────────────────────────────────────

_THAI = re.compile(r'[฀-๿]')
_PERIOD_SUFFIX = re.compile(
    r'_(?:BI-H[12]|Annual|annual|\d{4}\.\d[-–]\d{4}\.\d|\d{4}\.\d)(?:_\d+)?$',
    re.IGNORECASE,
)


def _thai_ratio(s: str) -> float:
    if not s:
        return 0.0
    return sum(1 for c in s if _THAI.match(c)) / len(s)


def _norm(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def title_en_from_report_name(name: str) -> str:
    """'9786161844684 - Je ne suis pas ta maman - ปิ๊บน้อยตามหาแม่.xlsx'
       → 'je ne suis pas ta maman'"""
    stem = Path(name).stem
    # strip ISBN prefix (10-13 digits at start)
    stem = re.sub(r'^\d{10,13}\s*-\s*', '', stem)
    # split by ' - ', take segments with low Thai ratio
    segs = stem.split(' - ')
    for seg in segs:
        if _thai_ratio(seg) < 0.3:
            return _norm(seg)
    return _norm(segs[0])


def title_en_from_legacy_name(name: str) -> str:
    """'Je ne suis pas ta maman - ปิ๊บน้อยตามหาแม่_BI-H1.xlsx'
       → 'je ne suis pas ta maman'"""
    stem = Path(name).stem
    stem = _PERIOD_SUFFIX.sub('', stem)
    # also strip trailing _2, _3 duplicate counters
    stem = re.sub(r'_\d+$', '', stem)
    segs = stem.split(' - ')
    return _norm(segs[0])


# ── แปลง report ที่ generate แล้ว → ข้อมูลย้อนหลังของปีนั้น ─────────────────────────

PERIOD_SUFFIX_BY_ID = {'bi1': 'BI-H1', 'bi2': 'BI-H2', 'annual': 'Annual'}


def report_arcname_to_legacy(arcname: str, year: int, period_suffix: str) -> Optional[str]:
    """'Agent/Publisher/9786161844684 - Title EN - ชื่อไทย.xlsx'
       → '2025/Agent/Publisher/Title EN - ชื่อไทย_Annual.xlsx'"""
    parts = Path(arcname).parts
    if len(parts) < 2 or not arcname.endswith('.xlsx'):
        return None
    stem = re.sub(r'^\d{10,13}\s*-\s*', '', Path(parts[-1]).stem)
    return '/'.join([str(year), *parts[:-1], f'{stem}_{period_suffix}.xlsx'])


def archive_report_as_legacy(report_zip: Path, output_zip: Path, year: int,
                             period: str, base_legacy_zip: Optional[Path] = None) -> dict:
    """สร้าง legacy ZIP ที่มีข้อมูลของ report ปีที่ระบุเพิ่มเข้าไป

    report ที่ระบบ generate เองในปีหนึ่ง จะกลายเป็น "ข้อมูลย้อนหลัง" ของปีถัดไป แต่
    โครงสร้างชื่อไฟล์ต่างกัน (report มี ISBN นำหน้า / legacy มีปีเป็นโฟลเดอร์บนสุดและ
    period ต่อท้าย) ฟังก์ชันนี้แปลงให้ตรงกับที่ `build_legacy_index()` อ่านได้

    base_legacy_zip: legacy pack เดิมที่จะเอามารวมด้วย (ไฟล์ปีเดียวกันไม่ถูกทับ —
    ถ้าชื่อชนกันจะเติม _2, _3 เหมือนที่ legacy_converter ทำ)
    """
    suffix = PERIOD_SUFFIX_BY_ID.get(period, 'Annual')
    copied = skipped = carried = 0
    used: set[str] = set()

    output_zip.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as out:
        if base_legacy_zip and Path(base_legacy_zip).exists():
            with zipfile.ZipFile(base_legacy_zip, 'r') as base:
                for name in base.namelist():
                    if name.endswith('/'):
                        continue
                    out.writestr(name, base.read(name))
                    used.add(name)
                    carried += 1

        with zipfile.ZipFile(report_zip, 'r') as rep:
            for name in sorted(rep.namelist()):
                target = report_arcname_to_legacy(name, year, suffix)
                if target is None:
                    skipped += 1
                    continue
                if target in used:
                    stem, ext = target[:-5], target[-5:]
                    n = 2
                    while f'{stem}_{n}{ext}' in used:
                        n += 1
                    target = f'{stem}_{n}{ext}'
                out.writestr(target, rep.read(name))
                used.add(target)
                copied += 1

    return {'year': year, 'period': suffix, 'added': copied,
            'carried_over': carried, 'skipped': skipped}


# ── Build index of legacy ZIP ───────────────────────────────────────────────────

_PERIOD_EXTRACT = re.compile(
    r'_(BI-H[12]|Annual|annual|\d{4}\.\d[-–]\d{4}\.\d|\d{4}\.\d)(?:_\d+)?$',
    re.IGNORECASE,
)


def build_legacy_index(legacy_zip_path: Path) -> dict[str, list[tuple[int, str, str]]]:
    """
    Returns {norm_title_en: [(year, period_label, zip_arcname), ...]}
    sorted by year asc, then period asc within each title.
    Deduplicates by (year, period) — keeps first occurrence only.
    """
    index: dict[str, list] = {}
    seen: set[tuple[str, int, str]] = set()  # (norm_title, year, period)

    with zipfile.ZipFile(legacy_zip_path, 'r') as zf:
        for arcname in sorted(zf.namelist()):  # sorted → _1 before _2
            if not arcname.endswith('.xlsx'):
                continue
            parts = Path(arcname).parts
            if len(parts) < 2:
                continue
            try:
                year = int(parts[0])
            except ValueError:
                continue
            fname = parts[-1]
            title = title_en_from_legacy_name(fname)
            if not title:
                continue
            # extract clean period label (without duplicate counter)
            m = _PERIOD_EXTRACT.search(Path(fname).stem)
            period_label = m.group(1) if m else 'Annual'
            period_label = period_label.upper().replace('ANNUAL', 'Annual')

            key = (title, year, period_label)
            if key in seen:
                continue  # skip duplicates
            seen.add(key)
            index.setdefault(title, []).append((year, period_label, arcname))

    # sort each entry by year then period
    for k in index:
        index[k].sort(key=lambda x: (x[0], x[1]))
    return index


# ── Excel cell copy ─────────────────────────────────────────────────────────────

def _copy_ws(ws_src, ws_dst, row_offset: int):
    """Copy all cells, styles, merges, and row heights from ws_src into ws_dst at row_offset."""
    # cell values + styles
    for row in ws_src.iter_rows():
        for cell in row:
            dst = ws_dst.cell(row=row_offset + cell.row - 1, column=cell.column)
            dst.value = cell.value
            if cell.has_style:
                dst.font       = copy(cell.font)
                dst.fill       = copy(cell.fill)
                dst.border     = copy(cell.border)
                dst.alignment  = copy(cell.alignment)
                dst.number_format = cell.number_format

    # merged cell ranges
    for mc in ws_src.merged_cells.ranges:
        ws_dst.merge_cells(
            start_row    = row_offset + mc.min_row - 1,
            start_column = mc.min_col,
            end_row      = row_offset + mc.max_row - 1,
            end_column   = mc.max_col,
        )

    # row heights
    for rn, rd in ws_src.row_dimensions.items():
        if rd.height:
            ws_dst.row_dimensions[row_offset + rn - 1].height = rd.height


def merge_excels(source_paths: list[Path], output_path: Path, gap_rows: int = 1):
    """Stack multiple Excel files vertically into one, preserving styles."""
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    current_row = 1
    col_widths_set = False

    for i, src in enumerate(source_paths):
        wb_src = openpyxl.load_workbook(src, data_only=True)
        ws_src = wb_src.active

        # copy column widths from first source only
        if not col_widths_set:
            for col_letter, cd in ws_src.column_dimensions.items():
                if cd.width:
                    ws_out.column_dimensions[col_letter].width = cd.width
            col_widths_set = True

        if i > 0:
            current_row += gap_rows  # blank separator row(s)

        _copy_ws(ws_src, ws_out, current_row)
        current_row += ws_src.max_row
        wb_src.close()

    wb_out.save(output_path)


# ── Main merge function ─────────────────────────────────────────────────────────

def merge_report_with_history(
    report_zip: Path,
    legacy_zip: Path,
    output_zip: Path,
) -> dict:
    """
    For each xlsx in report_zip, find matching legacy files by title_en,
    merge into one Excel (legacy years first, current year last).
    Unmatched report files are copied as-is.

    Returns stats dict: matched, unmatched, total
    """
    legacy_index = build_legacy_index(legacy_zip)

    tmp_dir    = Path(tempfile.mkdtemp())
    tmp_legacy = Path(tempfile.mkdtemp())
    tmp_out    = Path(tempfile.mkdtemp())

    stats = {'matched': 0, 'unmatched': 0, 'total': 0}

    try:
        # pre-extract legacy files we might need (deferred by title lookup)
        with zipfile.ZipFile(report_zip, 'r') as zf_rep, \
             zipfile.ZipFile(legacy_zip, 'r') as zf_leg:

            for arcname in zf_rep.namelist():
                if not arcname.endswith('.xlsx'):
                    continue
                stats['total'] += 1

                # extract current report file
                rep_path = tmp_dir / Path(arcname).name
                rep_path.parent.mkdir(parents=True, exist_ok=True)
                rep_data = zf_rep.read(arcname)
                rep_path.write_bytes(rep_data)

                # find matching legacy entries
                title = title_en_from_report_name(Path(arcname).name)
                matches = legacy_index.get(title, [])

                if not matches:
                    # no history → copy as-is
                    stats['unmatched'] += 1
                    out_path = tmp_out / arcname
                    out_path.parent.mkdir(parents=True, exist_ok=True)
                    out_path.write_bytes(rep_data)
                    continue

                # extract matching legacy files
                leg_paths = []
                for year, period_label, leg_arc in matches:
                    leg_data = zf_leg.read(leg_arc)
                    leg_local = tmp_legacy / f"{year}_{period_label}_{Path(leg_arc).name}"
                    leg_local.write_bytes(leg_data)
                    leg_paths.append(leg_local)

                # merge: legacy years (old→new) then current year
                out_path = tmp_out / arcname
                out_path.parent.mkdir(parents=True, exist_ok=True)
                merge_excels(leg_paths + [rep_path], out_path)
                stats['matched'] += 1

        # pack output ZIP
        with zipfile.ZipFile(output_zip, 'w', zipfile.ZIP_DEFLATED) as zf_out:
            for fpath in tmp_out.rglob('*.xlsx'):
                arcname = str(fpath.relative_to(tmp_out))
                zf_out.write(fpath, arcname)

    finally:
        shutil.rmtree(tmp_dir,    ignore_errors=True)
        shutil.rmtree(tmp_legacy, ignore_errors=True)
        shutil.rmtree(tmp_out,    ignore_errors=True)

    return stats
